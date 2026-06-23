#!/usr/bin/env python3
"""
update_time.py — Gala-day helper
Patches a swimmer's time in the club rankings HTML, updates the history block,
recalculates qualifying indicators, then commits and pushes to GitHub/Netlify.

Usage (shorthand for Margot or Ava):
  python3 update_time.py --margot "50 Back" SC 36.85 08/06/26 "Meet Name"
  python3 update_time.py --ava   "50 Free" SC 31.20 08/06/26 "Meet Name"

Usage (any swimmer in the table):
  python3 update_time.py "Bianca Sedda" "100 Back" SC 1:16.50 08/06/26 "Meet Name"

Arguments:
  swimmer   Full name as it appears in the table, OR --margot / --ava
  event     Event name: "50 Back", "100 Free", "200 Fly", etc.
  course    SC or LC
  time      Time string: 36.85 or 1:16.50
  date      DD/MM/YY  e.g. 08/06/26
  meet      Meet name in quotes (optional — defaults to "Gala")

Flags:
  --dry-run   Show what would change without writing files or pushing
  --no-push   Update files but don't commit/push
"""

import re
import sys
import json
import subprocess
import openpyxl
import datetime
import shutil
from pathlib import Path

# ── Paths ─────────────────────────────────────────────────────────────────────

GIT_ROOT          = Path("/Users/jamesmandikos/Documents/Claude Cowork/swim-results")
BROMPTON_DOCS_DIR = GIT_ROOT / "docs" / "brompton"
CWSC_DOCS_DIR     = GIT_ROOT / "docs" / "cwsc"
SWIM_ROOT         = GIT_ROOT / "Outputs" / "Swimming Results"

# (docs HTML,  xlsx,  qt_tab)
SWIMMER_FILES = {
    "Margot Mandikos": (
        BROMPTON_DOCS_DIR / "bsc_u14_girls_rankings.html",
        SWIM_ROOT / "Margot" / "MargotSwimTimes.xlsx",
        "2027-Times",
    ),
    "Ava Mandikos": (
        CWSC_DOCS_DIR / "cwsc_u14_girls_rankings.html",
        SWIM_ROOT / "Ava"    / "AvaSwimTimes.xlsx",
        "2026-Times",
    ),
}

# Peers live in Margot's file
PEER_FILE = BROMPTON_DOCS_DIR / "bsc_u14_girls_rankings.html"

# Manual swims files — persist times not yet on Swim England so regenerations don't lose them
MANUAL_SWIMS_FILES = {
    "Margot Mandikos": SWIM_ROOT / "Margot" / "margot_manual_swims.json",
    "Ava Mandikos":    SWIM_ROOT / "Ava"    / "ava_manual_swims.json",
}

# Event → stroke slug used in data-stroke attributes
STROKE_SLUG = {
    "50 Free":    "50-free",    "100 Free":   "100-free",  "200 Free":   "200-free",
    "400 Free":   "400-free",   "800 Free":   "800-free",  "1500 Free":  "1500-free",
    "50 Back":    "50-back",    "100 Back":   "100-back",  "200 Back":   "200-back",
    "50 Fly":     "50-fly",     "100 Fly":    "100-fly",   "200 Fly":    "200-fly",
    "50 Breast":  "50-breast",  "100 Breast": "100-breast","200 Breast": "200-breast",
    "100 IM":     "100-im",     "200 IM":     "200-im",    "400 IM":     "400-im",
}


# ── Time helpers ──────────────────────────────────────────────────────────────

def time_to_secs(t):
    t = str(t).strip()
    if not t or t == "-":
        return None
    try:
        if ":" in t:
            m, s = t.split(":")
            return int(m) * 60 + float(s)
        return float(t)
    except ValueError:
        return None


def secs_to_str(secs):
    if secs is None:
        return "-"
    secs = round(secs, 2)
    if secs >= 60:
        m = int(secs // 60)
        s = secs - m * 60
        return f"{m}:{s:05.2f}"
    return f"{secs:.2f}"


def _dt_secs(val):
    """Convert openpyxl time/timedelta value to seconds."""
    if val is None:
        return None
    if isinstance(val, datetime.time):
        return val.hour * 3600 + val.minute * 60 + val.second + val.microsecond / 1e6
    if isinstance(val, datetime.timedelta):
        return val.total_seconds()
    return None


# ── Read qualifying thresholds from spreadsheet ───────────────────────────────

def read_qt(xlsx_path, tab_name):
    """
    Returns dict: {(event, course): {county_qt, county_cons, region_qt, region_cons}}
    county_qual / region_qual = "Q", "C", or ""  (combined SC+LC best)
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[tab_name]
    raw = {}
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=False):
        c_val = row[1].value
        e_val = row[2].value
        if not c_val or not e_val:
            continue
        course = "SC" if str(c_val).strip() == "S" else "LC"
        event  = str(e_val).strip()
        r      = row[0].row
        raw[(event, course)] = {
            "pb_secs":    _dt_secs(ws.cell(r, 4).value),
            "county_qt":  _dt_secs(ws.cell(r, 7).value),
            "county_cons":_dt_secs(ws.cell(r, 9).value),
            "region_qt":  _dt_secs(ws.cell(r, 13).value),
            "region_cons":_dt_secs(ws.cell(r, 15).value),
        }

    # Combine SC+LC so either course can earn the qualification
    result = {}
    for (event, course), d in raw.items():
        def _status(pb, qt, cons):
            if pb and qt and pb <= qt:    return "Q"
            if pb and cons and pb <= cons: return "C"
            return ""
        paired = "LC" if course == "SC" else "SC"
        pd = raw.get((event, paired), {})
        def _combined(key):
            a = _status(d.get("pb_secs"), d.get(key), d.get(key.replace("_qt","_cons").replace("_cons","_qt")) if "qt" in key else None)
            # simpler: just check both directly
            return None
        # Compute per-course status then combine
        cq_this   = _status(d["pb_secs"], d["county_qt"], d["county_cons"])
        cq_paired = _status(pd.get("pb_secs"), pd.get("county_qt"), pd.get("county_cons"))
        rq_this   = _status(d["pb_secs"], d["region_qt"], d["region_cons"])
        rq_paired = _status(pd.get("pb_secs"), pd.get("region_qt"), pd.get("region_cons"))
        def best(a, b):
            if "Q" in (a, b): return "Q"
            if "C" in (a, b): return "C"
            return ""
        result[(event, course)] = {
            **d,
            "county_qual": best(cq_this, cq_paired),
            "region_qual": best(rq_this, rq_paired),
        }
    return result


def qual_status_for_time(new_secs, event, course, qt_data):
    """
    Given a new time (seconds), compute the county_qual and region_qual status
    by taking the BEST of (new time vs QTs) and (existing paired-course time vs QTs).
    Returns (county_qual, region_qual) each "Q", "C", or "".
    """
    d  = qt_data.get((event, course), {})
    paired = "LC" if course == "SC" else "SC"
    pd = qt_data.get((event, paired), {})

    def _s(pb, qt, cons):
        if pb and qt and pb <= qt:    return "Q"
        if pb and cons and pb <= cons: return "C"
        return ""

    def best(a, b):
        if "Q" in (a, b): return "Q"
        if "C" in (a, b): return "C"
        return ""

    cq_new    = _s(new_secs,           d.get("county_qt"),  d.get("county_cons"))
    cq_paired = _s(pd.get("pb_secs"), pd.get("county_qt"), pd.get("county_cons"))
    rq_new    = _s(new_secs,           d.get("region_qt"),  d.get("region_cons"))
    rq_paired = _s(pd.get("pb_secs"), pd.get("region_qt"), pd.get("region_cons"))

    return best(cq_new, cq_paired), best(rq_new, rq_paired)


def build_qual_ind(county_q, region_q):
    """Build the qual-ind HTML span from county/region qual status."""
    parts = []
    if county_q == "Q":
        parts.append('<span class="qi-q">CQ</span>')
    elif county_q == "C":
        parts.append('<span class="qi-c">CC</span>')
    if region_q == "Q":
        parts.append('<span class="qi-rq">RQ</span>')
    elif region_q == "C":
        parts.append('<span class="qi-rc">RC</span>')
    if not parts:
        return ""
    return '<span class="qual-ind">' + '<span class="qi-sep">·</span>'.join(parts) + '</span>'


# ── HTML patching ─────────────────────────────────────────────────────────────

def find_cell(html, swimmer_name, event, course):
    """
    Locate the event <td> for a given swimmer/event/course.
    Returns (match_object, is_empty) or (None, None).
    """
    slug = STROKE_SLUG.get(event)
    if not slug:
        return None, None
    ct = course.lower()
    esc = re.escape(swimmer_name)

    # Full cell (has data-secs)
    pat = (
        rf'<td class="event[^"]*" data-stroke="{slug}" data-coltype="{ct}"'
        rf' onclick="showHistory\(\'{esc}\''
        rf'.*?</td>'
    )
    m = re.search(pat, html, re.DOTALL)
    if m:
        is_empty = 'class="event empty"' in m.group(0) or '>—<' in m.group(0)
        return m, is_empty
    return None, None


def build_cell(swimmer_name, event, course, time_str, date_str, secs,
               qual_ind, old_secs, progress_arrow=""):
    """Reconstruct the full event <td> HTML."""
    slug = STROKE_SLUG[event]
    ct   = course.lower()
    title = f"Click to see {swimmer_name}'s {event} {course} history"
    return (
        f'<td class="event " data-stroke="{slug}" data-coltype="{ct}"'
        f' onclick="showHistory(\'{swimmer_name}\',\'{event}\',\'{course}\')"'
        f' style="cursor:pointer" title="{title}" data-secs="{secs:.3f}">'
        f'<span class="time">{time_str}</span>'
        f'{progress_arrow}'
        f'<br><span class="date">{date_str}</span>'
        f'{qual_ind}</td>'
    )


def update_history(html, swimmer_name, event, course, time_str, date_str, meet_name):
    """Insert a new entry at the top of the HISTORY block for this swimmer/event."""
    hist_match = re.search(r'var HISTORY=(\{.*?\});', html, re.DOTALL)
    if not hist_match:
        print("  WARNING: HISTORY block not found — skipping history update")
        return html

    history = json.loads(hist_match.group(1))
    key = f"{event}|{course}"
    swimmer_hist = history.setdefault(swimmer_name, {})
    event_list   = swimmer_hist.setdefault(key, [])

    # Don't duplicate if exact same time+date already exists
    for entry in event_list:
        if entry.get("time") == time_str and entry.get("date") == date_str:
            print(f"  History already contains {time_str} on {date_str} — skipping duplicate")
            return html

    new_entry = {"time": time_str, "pts": "", "round": "F", "date": date_str, "meet": meet_name}
    event_list.insert(0, new_entry)

    new_hist_json = json.dumps(history, separators=(",", ":"))
    html = html[:hist_match.start()] + f"var HISTORY={new_hist_json};" + html[hist_match.end():]
    return html


# ── Main update logic ─────────────────────────────────────────────────────────

def patch_html(html_path, swimmer_name, event, course, time_str, date_str, meet_name,
               qt_data, dry_run=False):
    """
    Patch the HTML file in-place. Returns True if a change was made.
    Always adds to history. Only updates the main cell when it's a new PB.
    """
    html = html_path.read_text(encoding="utf-8")

    new_secs = time_to_secs(time_str)
    if new_secs is None:
        print(f"  ERROR: Cannot parse time '{time_str}'")
        return False

    cell_match, is_empty = find_cell(html, swimmer_name, event, course)
    if cell_match is None:
        print(f"  WARNING: Cell for {swimmer_name} / {event} {course} not found in {html_path.name}")
        return False

    old_cell = cell_match.group(0)

    # Extract current time from data-secs attribute
    secs_m = re.search(r'data-secs="([\d.]+)"', old_cell)
    old_secs = float(secs_m.group(1)) if secs_m else None
    old_time_str = secs_to_str(old_secs) if old_secs else "—"
    is_pb = old_secs is None or new_secs < old_secs

    if dry_run:
        county_q, region_q = qual_status_for_time(new_secs, event, course, qt_data)
        print(f"\n  [DRY RUN] {swimmer_name} {event} {course}")
        print(f"    PB: {old_time_str}  →  {time_str} ({'new PB ▲' if is_pb else 'not PB — history only'})")
        print(f"    County: {county_q or '—'}  Regional: {region_q or '—'}")
        return True

    # Only update the main cell if this is a new PB
    if is_pb or is_empty:
        progress_arrow = ""
        if is_pb and old_secs is not None:
            improvement = old_secs - new_secs
            progress_arrow = (
                f'<span class="progress-arrow" title="PB set at most recent meet">'
                f' ▲<span style="font-size:9px;font-weight:normal;opacity:0.75"> -{improvement:.2f}s</span></span>'
            )
        elif is_pb:
            progress_arrow = '<span class="progress-arrow" title="PB set at most recent meet"> ▲</span>'

        county_q, region_q = qual_status_for_time(new_secs, event, course, qt_data)
        qual_ind = build_qual_ind(county_q, region_q)

        new_cell = build_cell(swimmer_name, event, course, time_str, date_str, new_secs,
                              qual_ind, old_secs, progress_arrow)
        html = html[:cell_match.start()] + new_cell + html[cell_match.end():]
        cell_note = f"{old_time_str} → {time_str} ▲ PB" + (f"  [{county_q} county]" if county_q else "") + (f"  [{region_q} regional]" if region_q else "")
    else:
        cell_note = f"PB stays {old_time_str} — {time_str} added to history only"

    # Always update history
    html = update_history(html, swimmer_name, event, course, time_str, date_str, meet_name)

    html_path.write_text(html, encoding="utf-8")
    print(f"  ✓ {swimmer_name} {event} {course}: {cell_note}")
    return True


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    args = sys.argv[1:]
    dry_run  = "--dry-run" in args;  args = [a for a in args if a != "--dry-run"]
    no_push  = "--no-push" in args;  args = [a for a in args if a != "--no-push"]

    if len(args) < 5:
        print(__doc__)
        sys.exit(1)

    # Resolve swimmer name
    if args[0] == "--margot":
        swimmer_name = "Margot Mandikos"
        args = args[1:]
    elif args[0] == "--ava":
        swimmer_name = "Ava Mandikos"
        args = args[1:]
    else:
        swimmer_name = args[0]
        args = args[1:]

    if len(args) < 4:
        print("Not enough arguments."); print(__doc__); sys.exit(1)

    event, course, time_str, date_str = args[0], args[1], args[2], args[3]
    meet_name = args[4] if len(args) > 4 else "Gala"
    course = course.upper()

    if course not in ("SC", "LC"):
        print(f"Course must be SC or LC, got: {course}"); sys.exit(1)
    if event not in STROKE_SLUG:
        print(f"Unknown event '{event}'. Valid events: {', '.join(STROKE_SLUG)}")
        sys.exit(1)

    print(f"\nUpdate: {swimmer_name} | {event} {course} | {time_str} | {date_str} | {meet_name}")
    print("-" * 60)

    # Determine which HTML files to update and which QT spreadsheet to use
    if swimmer_name in SWIMMER_FILES:
        docs_html, xlsx, qt_tab = SWIMMER_FILES[swimmer_name]
    else:
        # Peer swimmer — lives in Margot's file; use Margot's QTs as best approximation
        docs_html = PEER_FILE
        _, xlsx, qt_tab = SWIMMER_FILES["Margot Mandikos"]

    # Load qualifying thresholds
    try:
        qt_data = read_qt(xlsx, qt_tab)
    except Exception as e:
        print(f"  WARNING: Could not load QT data ({e}) — qual indicators will be blank")
        qt_data = {}

    changed = patch_html(docs_html, swimmer_name, event, course,
                         time_str, date_str, meet_name, qt_data, dry_run)

    # Persist to manual swims JSON so future regenerations from swimming_results.py
    # don't lose times that haven't appeared on Swim England yet
    manual_path = MANUAL_SWIMS_FILES.get(swimmer_name)
    if changed and not dry_run and manual_path:
        try:
            manual = json.loads(manual_path.read_text()) if manual_path.exists() else {}
            key = f"{event}|{course}"
            entries = manual.setdefault(key, [])
            if not any(e["date"] == date_str and e["time"] == time_str for e in entries):
                entries.append({"time": time_str, "pts": "", "round": "F",
                                 "date": date_str, "meet": meet_name})
                manual_path.write_text(json.dumps(manual, indent=2))
                print(f"  ✓ Saved to {manual_path.name}")
        except Exception as e:
            print(f"  WARNING: could not update manual swims file: {e}")

    if not changed or dry_run or no_push:
        if dry_run: print("\n  [DRY RUN] No files written.")
        return

    # Commit and push
    print("\nCommitting and pushing...")
    rel = str(docs_html.relative_to(GIT_ROOT))
    msg = f"Gala update: {swimmer_name} {event} {course} {time_str} ({date_str})"
    try:
        subprocess.run(["git", "-C", str(GIT_ROOT), "add", rel], check=True)
        subprocess.run(["git", "-C", str(GIT_ROOT), "commit", "-m", msg], check=True)
        subprocess.run(["git", "-C", str(GIT_ROOT), "push"], check=True)
        print(f"  ✓ Pushed — Netlify will deploy in ~30 seconds")
    except subprocess.CalledProcessError as e:
        print(f"  ERROR during git: {e}")
        print("  File was updated locally — run 'git push' manually when ready.")


if __name__ == "__main__":
    main()
