"""
Update Qualifying Times — swimmingresults.org → MargotSwimTimes.xlsx
Fetches Middlesex County and London Regional qualifying times for all age groups
(female, SC and LC) and writes them into the spreadsheet.

Also builds/updates the 2027-Times tab with age-13 QTs.

Run this once a year when new qualifying times are published.
Usage: python3 update_qualifying_times.py
"""

import subprocess
import tempfile
import os
import re
import datetime
from pathlib import Path

from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPTS_DIR = Path(__file__).parent
ROOT_DIR    = SCRIPTS_DIR.parent
XLSX_PATH   = ROOT_DIR / "Margot" / "MargotSwimTimes.xlsx"

TIME_FORMAT = "m:ss.00"  # Excel number format for swim times (MM:SS.hundredths)


def to_time(t):
    """Convert '1:07.90' or '30.96' string to datetime.time. Returns None for '-'/'NONE'."""
    if not t or t in ("-", "NONE"):
        return None
    t = str(t).strip()
    try:
        if ":" in t:
            m, s = t.split(":")
            total = int(m) * 60 + float(s)
        else:
            total = float(t)
        mins    = int(total // 60)
        secs    = total % 60
        sec_int = int(secs)
        us      = round((secs - sec_int) * 1_000_000)
        return datetime.time(0, mins, sec_int, us)
    except (ValueError, AttributeError):
        return None

UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)

# Events in order as they appear in the QT table (column index matches)
QT_EVENTS = [
    "50m Fr", "100m Fr", "200m Fr", "400m Fr", "800m Fr", "1500m Fr",
    "50m Br", "100m Br", "200m Br",
    "50m Fly", "100m Fly", "200m Fly",
    "50m Ba", "100m Ba", "200m Ba",
    "200m IM", "400m IM",
]

# Map QT event names → spreadsheet event names (as used in 2026-Times col C)
QT_TO_SHEET = {
    "50m Fr":   "50 Free",   "100m Fr":  "100 Free",  "200m Fr":  "200 Free",
    "400m Fr":  "400 Free",  "800m Fr":  "800 Free",  "1500m Fr": "1500 Free",
    "50m Br":   "50 Breast", "100m Br":  "100 Breast","200m Br":  "200 Breast",
    "50m Fly":  "50 Fly",    "100m Fly": "100 Fly",   "200m Fly": "200 Fly",
    "50m Ba":   "50 Back",   "100m Ba":  "100 Back",  "200m Ba":  "200 Back",
    "200m IM":  "200 IM",    "400m IM":  "400 IM",
}

AGE_GROUPS = ["09","10","11","12","13","14","15","16","17","18"]


# ── HTTP ──────────────────────────────────────────────────────────────────────

def fetch_qt_page(pool, level_code, region=None, county=None, year=2026):
    """Fetch the qualifying times generator page and return BeautifulSoup."""
    with tempfile.NamedTemporaryFile(suffix=".html", delete=False) as f:
        tmp = f.name

    strokes = [1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17]
    stroke_qs = "&".join(f"Stroke[]={s}" for s in strokes)
    age_qs    = "&".join(f"AgeGroup[]={a}" for a in AGE_GROUPS)
    region_p  = f"TargetRegion={region}" if region else "TargetRegion=P"
    county_p  = f"TargetCounty={county}" if county else "TargetCounty=XXXX"
    date_end  = f"31%2F12%2F{year}"
    date_start = f"01%2F01%2F{year}"

    params = (
        f"Pool={pool}&Sex=F&{stroke_qs}&{age_qs}"
        f"&date={date_end}&date_start={date_start}&date_end={date_end}"
        f"&PosFina=P&Number=25&StartNumber=1"
        f"&Level={level_code}&TargetNationality=X&{region_p}&{county_p}"
    )
    url = f"https://www.swimmingresults.org/qt/qtmake.php?{params}"

    subprocess.run([
        "curl", "-s", "--max-time", "20", "--compressed",
        "-H", f"User-Agent: {UA}",
        "-H", "Referer: https://www.swimmingresults.org/qt/",
        "-o", tmp, url
    ], check=False)

    with open(tmp, encoding="utf-8", errors="replace") as f:
        html = f.read()
    os.unlink(tmp)
    return BeautifulSoup(html, "html.parser")


def parse_qt_table(soup):
    """
    Parse table 3 (the combined data table) from the QT page.
    Returns dict: {(age_int, "Lower"|"Upper"): {event_name: time_str}}
    """
    tables = soup.find_all("table")
    if len(tables) < 3:
        return {}
    table = tables[2]
    rows  = table.find_all("tr")
    data  = {}

    for row in rows[1:]:  # skip header
        cells = [td.get_text(strip=True) for td in row.find_all(["th", "td"])]
        if len(cells) < 3:
            continue
        age_str  = cells[0].strip()
        row_type = cells[1].strip()  # "Lower" or "Upper"
        if not age_str.isdigit():
            continue
        age = int(age_str)
        times = {}
        for i, event in enumerate(QT_EVENTS):
            val = cells[2 + i] if (2 + i) < len(cells) else "NONE"
            times[event] = val if val != "NONE" else "-"
        data[(age, row_type)] = times

    return data


def fetch_all_qt(year=2026):
    """
    Fetch all QTs for Middlesex County and London Regional, SC and LC.
    Returns: {
        "County SC":   {(age, type): {event: time}},
        "County LC":   ...,
        "Regional SC": ...,
        "Regional LC": ...,
    }
    """
    result = {}
    configs = [
        ("County SC",   "S", "C", None,  "MDXL"),
        ("County LC",   "L", "C", None,  "MDXL"),
        ("Regional SC", "S", "D", "L",   None),
        ("Regional LC", "L", "D", "L",   None),
    ]
    for label, pool, level, region, county in configs:
        print(f"  Fetching {label}...")
        soup = fetch_qt_page(pool, level, region=region, county=county, year=year)
        data = parse_qt_table(soup)
        result[label] = data
        print(f"    → {len(data)} age/type rows")
    return result


# ── Excel helpers ─────────────────────────────────────────────────────────────

def make_styles():
    return {
        "hdr_fill":    PatternFill("solid", fgColor="1F4E79"),
        "lower_fill":  PatternFill("solid", fgColor="E8F5E9"),
        "upper_fill":  PatternFill("solid", fgColor="FFF8E1"),
        "age_fill":    PatternFill("solid", fgColor="EBF5FB"),
        "sc_fill":     PatternFill("solid", fgColor="D6EAF8"),
        "lc_fill":     PatternFill("solid", fgColor="D5F5E3"),
        "hdr_font":    Font(color="FFFFFF", bold=True, size=9),
        "bold_font":   Font(bold=True, size=9),
        "normal_font": Font(size=9),
        "center":      Alignment(horizontal="center", vertical="center", wrap_text=True),
        "left":        Alignment(horizontal="left", vertical="center"),
        "thin":        Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        ),
    }


def write_qt_sheet(wb, sheet_name, sc_data, lc_data, year):
    """Write a qualifying times sheet showing SC and LC side by side."""
    st = make_styles()

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    events = QT_EVENTS  # 17 events
    # Layout: Age | SC Lower | SC Upper | LC Lower | LC Upper  per event
    # Group: one pair of SC/LC columns per event

    # Row 1: title
    ws.cell(row=1, column=1, value=f"{sheet_name} — Female — All Age Groups")
    ws.cell(row=1, column=1).font = Font(bold=True, size=11, color="1F4E79")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(events)*4)

    # Row 2: event group headers (span 4 cols each: SC Lower, SC Upper, LC Lower, LC Upper)
    col = 2
    ws.cell(row=2, column=1, value="Age").fill = st["hdr_fill"]
    ws.cell(row=2, column=1).font = st["hdr_font"]
    ws.cell(row=2, column=1).alignment = st["center"]
    ws.cell(row=2, column=1).border = st["thin"]

    for event in events:
        for c in range(col, col+4):
            ws.cell(row=2, column=c).fill = st["hdr_fill"]
            ws.cell(row=2, column=c).border = st["thin"]
        ws.cell(row=2, column=col, value=event)
        ws.cell(row=2, column=col).font = st["hdr_font"]
        ws.cell(row=2, column=col).alignment = st["center"]
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+3)
        col += 4

    # Row 3: SC Lower | SC Upper | LC Lower | LC Upper sub-headers
    ws.cell(row=3, column=1, value="").border = st["thin"]
    col = 2
    for _ in events:
        labels = [("SC QT", st["sc_fill"]), ("SC Cons", st["sc_fill"]),
                  ("LC QT", st["lc_fill"]), ("LC Cons", st["lc_fill"])]
        for label, fill in labels:
            c = ws.cell(row=3, column=col, value=label)
            c.fill  = fill; c.font = Font(bold=True, size=8)
            c.alignment = st["center"]; c.border = st["thin"]
            col += 1

    # Data rows — one row per age group
    row_num = 4
    for age in range(9, 19):
        ws.cell(row=row_num, column=1, value=f"{age} yrs")
        ws.cell(row=row_num, column=1).fill = st["age_fill"]
        ws.cell(row=row_num, column=1).font = st["bold_font"]
        ws.cell(row=row_num, column=1).alignment = st["center"]
        ws.cell(row=row_num, column=1).border = st["thin"]

        col = 2
        for event in events:
            sc_lower = sc_data.get((age, "Lower"), {}).get(event, "-")
            sc_upper = sc_data.get((age, "Upper"), {}).get(event, "-")
            lc_lower = lc_data.get((age, "Lower"), {}).get(event, "-")
            lc_upper = lc_data.get((age, "Upper"), {}).get(event, "-")

            for val, fill in [
                (sc_lower, st["sc_fill"]), (sc_upper, st["sc_fill"]),
                (lc_lower, st["lc_fill"]), (lc_upper, st["lc_fill"])
            ]:
                t_val = to_time(val)
                c = ws.cell(row=row_num, column=col, value=t_val)
                c.number_format = TIME_FORMAT
                c.fill      = fill if t_val is not None else PatternFill()
                c.font      = st["normal_font"]
                c.alignment = st["center"]
                c.border    = st["thin"]
                col += 1

        row_num += 1

    # Column widths
    ws.column_dimensions["A"].width = 8
    for i in range(len(events) * 4):
        ws.column_dimensions[get_column_letter(2+i)].width = 9

    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 16
    ws.freeze_panes = "B4"

    print(f"  → {sheet_name} written")


# ── Build 2027-Times tab ──────────────────────────────────────────────────────

def build_2027_times(wb, all_qt):
    """
    Create 2027-Times tab by copying 2026-Times structure and
    replacing County QT, County Cons, Regional QT, Regional Cons
    columns with age-13 times (Margot's age group in 2027).
    """
    TARGET_AGE = 13
    YEAR       = 2027

    if "2027-Times" in wb.sheetnames:
        del wb["2027-Times"]

    # Copy 2026-Times as a base
    src = wb["2026-Times"]
    ws  = wb.copy_worksheet(src)
    ws.title = "2027-Times"

    # Find column indices from row 3 headers
    header_row = 3
    col_map = {}  # header text → column index
    for cell in ws[header_row]:
        if cell.value:
            col_map[str(cell.value).strip()] = cell.column

    # Columns to update (using header text from 2026-Times row 3)
    # County QT → col G, County Cons → col N, Regional QT → col P (approx — find by header)
    county_qt_cols  = []
    county_con_cols = []
    region_qt_cols  = []
    region_con_cols = []

    for cell in ws[header_row]:
        v = str(cell.value or "").strip()
        if "County QT" in v or v == "County QT":
            county_qt_cols.append(cell.column)
        elif "Cons" in v and "Regional" not in v:
            county_con_cols.append(cell.column)
        elif "Regional" in v and "QT" in v and "Cons" not in v:
            region_qt_cols.append(cell.column)
        elif "Regional" in v and "Cons" in v:
            region_con_cols.append(cell.column)

    # Map spreadsheet event name + course → age-13 QT time
    def get_qt(sheet_event, course_letter, qt_type, qt_key):
        """qt_key: "County SC", "County LC", "Regional SC", "Regional LC" """
        data = all_qt.get(qt_key, {})
        times = data.get((TARGET_AGE, qt_type), {})
        # Find matching QT event
        for qt_event, sheet_name in QT_TO_SHEET.items():
            if sheet_name == sheet_event:
                return times.get(qt_event, "-")
        return "-"

    # Update data rows (row 4 onwards)
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        course_cell = row[1]   # col B
        event_cell  = row[2]   # col C
        if not course_cell.value or not event_cell.value:
            continue
        course = str(course_cell.value).strip()  # "S" or "L"
        event  = str(event_cell.value).strip()
        course_suffix = "SC" if course == "S" else "LC"

        county_qt  = get_qt(event, course, "Lower", f"County {course_suffix}")
        county_con = get_qt(event, course, "Upper", f"County {course_suffix}")
        region_qt  = get_qt(event, course, "Lower", f"Regional {course_suffix}")
        region_con = get_qt(event, course, "Upper", f"Regional {course_suffix}")

        if event == "100 IM":
            continue  # not a county or regional event

        for col, val in [
            *[(col, county_qt)  for col in county_qt_cols],
            *[(col, county_con) for col in county_con_cols],
            *[(col, region_qt)  for col in region_qt_cols],
            *[(col, region_con) for col in region_con_cols],
        ]:
            c = ws.cell(row=row[0].row, column=col, value=to_time(val))
            c.number_format = TIME_FORMAT

    # Clear Margot's PBs, dates and rankings (she hasn't swum these times yet)
    pb_col   = col_map.get("PB", 4)
    date_col = col_map.get("Obtained", 5)
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        if row[1].value:  # has a course
            ws.cell(row=row[0].row, column=pb_col, value=None)
            ws.cell(row=row[0].row, column=date_col, value=None)

    # Update title cell if present
    for cell in ws[1]:
        if cell.value and "2026" in str(cell.value):
            cell.value = str(cell.value).replace("2026", "2027")

    print(f"  → 2027-Times tab written (age-{TARGET_AGE} QTs)")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print(f"Update Qualifying Times — {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print("=" * 55)

    print("\nFetching qualifying times from swimmingresults.org...")
    all_qt = fetch_all_qt(year=2026)

    print("\nUpdating spreadsheet...")
    wb = openpyxl.load_workbook(XLSX_PATH)

    print("\nBuilding Regional Times 2026 tab...")
    write_qt_sheet(
        wb, "Regional Times 2026",
        sc_data=all_qt["Regional SC"],
        lc_data=all_qt["Regional LC"],
        year=2026
    )

    print("\nBuilding County Times 2026 tab...")
    write_qt_sheet(
        wb, "County Times 2026",
        sc_data=all_qt["County SC"],
        lc_data=all_qt["County LC"],
        year=2026
    )

    print("\nBuilding 2027-Times tab...")
    build_2027_times(wb, all_qt)

    wb.save(XLSX_PATH)
    print(f"\nDone — saved to {XLSX_PATH.name}")


if __name__ == "__main__":
    main()
