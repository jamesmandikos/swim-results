"""
Swimming Results — Margot Mandikos
Fetches personal best times for all swimmers in swimmers_config.json
using the individual best times endpoint (no Cloudflare issues).
Outputs:
  1. Updated PBs in MargotSwimTimes.xlsx → 2026-Times tab (only if faster)
  2. club_rankings.html — browser report with colour coding

To add a new swimmer: edit swimmers_config.json and add their name and member ID.
Member IDs can be found at swimmingresults.org/individualbest/ by searching their surname.
"""

import json
import subprocess
import tempfile
import os
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import urlencode

from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────

SCRIPTS_DIR  = Path(__file__).parent
ROOT_DIR     = SCRIPTS_DIR.parent
MARGOT_DIR   = ROOT_DIR / "Margot"
AVA_DIR      = ROOT_DIR / "Ava"
BROMPTON_DOCS_DIR = ROOT_DIR.parent.parent / "docs" / "brompton"
CWSC_DOCS_DIR     = ROOT_DIR.parent.parent / "docs" / "cwsc"

CONFIG_PATH          = MARGOT_DIR / "margot_swimmers.json"
CONFIG_2013_PATH     = MARGOT_DIR / "margot_swimmers_2013.json"
CONFIG_2015_PATH     = MARGOT_DIR / "margot_swimmers_2015.json"
XLSX_PATH            = MARGOT_DIR / "MargotSwimTimes.xlsx"
HTML_PATH            = BROMPTON_DOCS_DIR / "bsc_u14_girls_rankings.html"
PEER_HISTORY_PATH    = MARGOT_DIR / "peer_histories.json"
PEER_HIST_2013_PATH  = MARGOT_DIR / "peer_histories_2013.json"
PEER_HIST_2015_PATH  = MARGOT_DIR / "peer_histories_2015.json"
GALAS_PATH           = ROOT_DIR  / "galas.json"
AVA_XLSX_PATH            = AVA_DIR    / "AvaSwimTimes.xlsx"
AVA_HTML_PATH            = CWSC_DOCS_DIR / "cwsc_u14_girls_rankings.html"
AVA_CONFIG_PATH          = AVA_DIR    / "ava_swimmers.json"
AVA_PEER_HIST_PATH       = AVA_DIR    / "ava_peer_histories.json"
AVA_CONFIG_2016_PATH     = AVA_DIR    / "ava_swimmers_2016.json"
AVA_PEER_HIST_2016_PATH  = AVA_DIR    / "ava_peer_histories_2016.json"
AVA_CONFIG_2018_PATH     = AVA_DIR    / "ava_swimmers_2018.json"
AVA_PEER_HIST_2018_PATH  = AVA_DIR    / "ava_peer_histories_2018.json"
CONFIG_2016_PATH         = MARGOT_DIR / "margot_swimmers_2016.json"
CONFIG_2017_PATH         = MARGOT_DIR / "margot_swimmers_2017.json"
CONFIG_2018_PATH         = MARGOT_DIR / "margot_swimmers_2018.json"
PEER_HIST_2016_PATH      = MARGOT_DIR / "peer_histories_2016.json"
PEER_HIST_2017_PATH      = MARGOT_DIR / "peer_histories_2017.json"
PEER_HIST_2018_PATH      = MARGOT_DIR / "peer_histories_2018.json"
AVA_CONFIG_2013_PATH     = AVA_DIR    / "ava_swimmers_2013.json"
AVA_CONFIG_2014_PATH     = AVA_DIR    / "ava_swimmers_2014.json"
AVA_CONFIG_2015_PATH     = AVA_DIR    / "ava_swimmers_2015.json"
AVA_PEER_HIST_2013_PATH  = AVA_DIR    / "ava_peer_histories_2013.json"
AVA_PEER_HIST_2014_PATH  = AVA_DIR    / "ava_peer_histories_2014.json"
AVA_PEER_HIST_2015_PATH  = AVA_DIR    / "ava_peer_histories_2015.json"

MARGOT_NAME  = "Margot Mandikos"
MARGOT_ID    = "1649626"
AVA_NAME     = "Ava Mandikos"
RATE_DELAY   = 2   # seconds between personal best fetches

# Map short event name → stroke code used by swimmingresults.org
STROKE_CODES = {
    "50 Free":"1","100 Free":"2","200 Free":"3","400 Free":"4",
    "800 Free":"5","1500 Free":"6",
    "50 Breast":"7","100 Breast":"8","200 Breast":"9",
    "50 Fly":"10","100 Fly":"11","200 Fly":"12",
    "50 Back":"13","100 Back":"14","200 Back":"15",
    "100 IM":"18","200 IM":"16","400 IM":"17",
}

UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)

# ── Event definitions ─────────────────────────────────────────────────────────

# Display order: SC | LC | LC→SC converted, per stroke
# (short_name, course)  course = "SC", "LC", or "LC→SC"
EVENTS = [
    ("50 Back",    "SC"), ("50 Back",    "LC"), ("50 Back",    "LC→SC"),
    ("100 Back",   "SC"), ("100 Back",   "LC"), ("100 Back",   "LC→SC"),
    ("200 Back",   "SC"), ("200 Back",   "LC"), ("200 Back",   "LC→SC"),
    ("50 Fly",     "SC"), ("50 Fly",     "LC"), ("50 Fly",     "LC→SC"),
    ("100 Fly",    "SC"), ("100 Fly",    "LC"), ("100 Fly",    "LC→SC"),
    ("200 Fly",    "SC"), ("200 Fly",    "LC"), ("200 Fly",    "LC→SC"),
    ("50 Free",    "SC"), ("50 Free",    "LC"), ("50 Free",    "LC→SC"),
    ("100 Free",   "SC"), ("100 Free",   "LC"), ("100 Free",   "LC→SC"),
    ("200 Free",   "SC"), ("200 Free",   "LC"), ("200 Free",   "LC→SC"),
    ("400 Free",   "SC"), ("400 Free",   "LC"), ("400 Free",   "LC→SC"),
    ("800 Free",   "SC"), ("800 Free",   "LC"), ("800 Free",   "LC→SC"),
    ("50 Breast",  "SC"), ("50 Breast",  "LC"), ("50 Breast",  "LC→SC"),
    ("100 Breast", "SC"), ("100 Breast", "LC"), ("100 Breast", "LC→SC"),
    ("200 Breast", "SC"), ("200 Breast", "LC"), ("200 Breast", "LC→SC"),
    ("100 IM",     "SC"), ("100 IM",     "LC"), ("100 IM",     "LC→SC"),
    ("200 IM",     "SC"), ("200 IM",     "LC"), ("200 IM",     "LC→SC"),
    ("400 IM",     "SC"), ("400 IM",     "LC"), ("400 IM",     "LC→SC"),
]

# Map site event names → our short names
SITE_EVENT_MAP = {
    "50 Freestyle":          "50 Free",
    "100 Freestyle":         "100 Free",
    "200 Freestyle":         "200 Free",
    "400 Freestyle":         "400 Free",
    "800 Freestyle":         "800 Free",
    "1500 Freestyle":        "1500 Free",
    "50 Backstroke":         "50 Back",
    "100 Backstroke":        "100 Back",
    "200 Backstroke":        "200 Back",
    "50 Breaststroke":       "50 Breast",
    "100 Breaststroke":      "100 Breast",
    "200 Breaststroke":      "200 Breast",
    "50 Butterfly":          "50 Fly",
    "100 Butterfly":         "100 Fly",
    "200 Butterfly":         "200 Fly",
    "100 Individual Medley": "100 IM",
    "200 Individual Medley": "200 IM",
    "400 Individual Medley": "400 IM",
}

# ── HTTP ──────────────────────────────────────────────────────────────────────

def curl_get(url, referer="https://www.swimmingresults.org/individualbest/"):
    with tempfile.NamedTemporaryFile(suffix=".html", delete=False) as f:
        tmp = f.name
    try:
        subprocess.run(
            ["curl", "-s", "--max-time", "30", "--compressed",
             "-H", f"User-Agent: {UA}",
             "-H", "Accept: text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
             "-H", "Accept-Language: en-GB,en;q=0.9",
             "-H", f"Referer: {referer}",
             "-o", tmp, url],
            check=False,
        )
        with open(tmp, encoding="utf-8", errors="replace") as f:
            html = f.read()
    finally:
        os.unlink(tmp)
    return BeautifulSoup(html, "html.parser")


# ── Time helpers ──────────────────────────────────────────────────────────────

def time_to_seconds(t):
    if not t or str(t).strip() in ("-", ""):
        return None
    t = str(t).strip()
    try:
        if ":" in t:
            m, s = t.split(":")
            return int(m) * 60 + float(s)
        return float(t)
    except ValueError:
        return None


def cell_to_seconds(val):
    import datetime as dt
    if val is None or val == "-":
        return None
    if isinstance(val, dt.time):
        return val.hour * 3600 + val.minute * 60 + val.second + val.microsecond / 1e6
    if isinstance(val, dt.timedelta):
        return val.total_seconds()
    if isinstance(val, dt.datetime):
        return val.hour * 3600 + val.minute * 60 + val.second + val.microsecond / 1e6
    return time_to_seconds(str(val))


def seconds_to_str(secs):
    if secs is None:
        return "-"
    secs = round(secs, 2)
    if secs >= 60:
        m = int(secs // 60)
        s = secs - m * 60
        return f"{m}:{s:05.2f}"
    return f"{secs:.2f}"


def ordinal(n):
    if 11 <= (n % 100) <= 13:
        return f"{n}th"
    return f"{n}{['th','st','nd','rd','th'][min(n % 10, 4)]}"


# ── Fetch swimmer data ────────────────────────────────────────────────────────

def fetch_swimmer_bests(member_id, name):
    """
    Fetch all personal bests (LC + SC) for a swimmer by member ID.
    Returns dict: {(short_name, course): {time_str, time_secs, date_str}}
    """
    url = (
        f"https://www.swimmingresults.org/individualbest/personal_best.php"
        f"?mode=A&tiref={member_id}"
    )
    soup = curl_get(url)
    bests = {}
    tables = soup.find_all("table", id="rankTable")
    for table in tables:
        header = table.find("tr")
        header_text = header.get_text(" ") if header else ""
        course = "SC" if "SC Time" in header_text else "LC"
        for row in table.find_all("tr")[1:]:
            cells = row.find_all("td")
            if len(cells) < 5:
                continue
            site_event = cells[0].get_text(strip=True)
            time_str   = cells[1].get_text(strip=True)
            converted  = cells[2].get_text(strip=True)  # equivalent time in other course
            pb_date    = cells[4].get_text(strip=True)
            short_name = SITE_EVENT_MAP.get(site_event)
            if not short_name:
                continue
            secs = time_to_seconds(time_str)
            if secs:
                bests[(short_name, course)] = {
                    "time_str":  time_str,
                    "time_secs": secs,
                    "date_str":  pb_date,
                    "converted": converted,
                }
    return bests


def fetch_margot_history():
    """
    Fetch all recorded swims for Margot for every stroke and course.
    Returns dict: {(short_name, "SC"|"LC"): [{time, pts, round, date, meet}]}
    """
    history = {}
    stroke_names = list(dict.fromkeys(n for n, c in EVENTS if c != "LC→SC"))
    courses = ["SC", "LC"]
    seen = set()
    for short_name, course in [(n, c) for n, c in EVENTS if c != "LC→SC"]:
        if (short_name, course) in seen:
            continue
        seen.add((short_name, course))
        stroke_code = STROKE_CODES.get(short_name)
        if not stroke_code:
            continue
        pool = "S" if course == "SC" else "L"
        url = (
            f"https://www.swimmingresults.org/individualbest/personal_best_time_date.php"
            f"?tiref={MARGOT_ID}&mode=A&tstroke={stroke_code}&tcourse={pool}"
            f"&Pool={pool}&Stroke={stroke_code}"
        )
        soup = curl_get(url, referer="https://www.swimmingresults.org/individualbest/")
        table = soup.find("table", id="rankTable")
        rows = []
        if table:
            for row in table.find_all("tr")[1:]:
                cells = [td.get_text(strip=True) for td in row.find_all("td")]
                if len(cells) >= 4:
                    rows.append({
                        "time":  cells[0],
                        "pts":   cells[1],
                        "round": cells[2],
                        "date":  cells[3],
                        "meet":  cells[4] if len(cells) > 4 else "",
                    })
        history[(short_name, course)] = rows
    return history


def fetch_all_swimmers():
    """
    Load config and fetch PBs for every swimmer.
    Returns list of {name, member_id, is_margot, bests}
    """
    config = json.loads(CONFIG_PATH.read_text())
    swimmers = config["swimmers"]
    yob = config.get("year_of_birth", 2014)
    results = []
    total = len(swimmers)
    for i, s in enumerate(swimmers, 1):
        print(f"  [{i}/{total}] {s['name']}...")
        bests = fetch_swimmer_bests(s["member_id"], s["name"])
        print(f"    → {len(bests)} times")
        results.append({
            "name":          s["name"],
            "member_id":     s["member_id"],
            "is_margot":     s["name"] == MARGOT_NAME,
            "year_of_birth": yob,
            "has_history":   True,
            "bests":         bests,
        })
        if i < total:
            time.sleep(RATE_DELAY)
    return results


# ── Build comparison matrix ───────────────────────────────────────────────────

def build_matrix(swimmers):
    """
    For each event, rank all swimmers by time (fastest first).
    Returns list of swimmer rows sorted by average WA points descending.
    Each row: {name, is_margot, avg_pts, events: {(name,course): {time_str, date_str, club_rank}}}
    """
    # Collect per-event times across all swimmers
    event_times = {}  # (short_name, course) → [(time_secs, swimmer_idx)]
    for idx, sw in enumerate(swimmers):
        for key, data in sw["bests"].items():
            event_times.setdefault(key, []).append((data["time_secs"], idx))

    # Sort each event (SC and LC only) by time, assign ranks
    event_ranks = {}  # (key, swimmer_idx) → rank
    for key, entries in event_times.items():
        entries.sort(key=lambda x: x[0])
        for rank, (_, idx) in enumerate(entries, 1):
            event_ranks[(key, idx)] = rank

    rows = []
    for idx, sw in enumerate(swimmers):
        events_out = {}
        rank_sum = 0
        rank_count = 0

        for key, data in sw["bests"].items():
            if key not in event_times:
                continue
            rank = event_ranks.get((key, idx), len(event_times[key]) + 1)
            events_out[key] = {
                "time_str":  data["time_str"],
                "date_str":  data["date_str"],
                "club_rank": rank,
            }
            rank_sum += rank
            rank_count += 1

            # Add LC→SC converted column from the LC entry
            if key[1] == "LC":
                converted = data.get("converted", "")
                conv_key = (key[0], "LC→SC")
                if converted and converted != "-":
                    events_out[conv_key] = {
                        "time_str":  converted,
                        "date_str":  data["date_str"],
                        "club_rank": None,  # ranked separately below
                    }

        events_out = dict(events_out)  # copy before ranking converted

        avg_rank = rank_sum / rank_count if rank_count else 999
        rows.append({
            "name":          sw["name"],
            "is_margot":     sw["is_margot"],
            "year_of_birth": sw.get("year_of_birth", 2014),
            "has_history":   sw.get("has_history", True),
            "avg_rank":      avg_rank,
            "event_count":   rank_count,
            "events":        events_out,
        })

    # Rank the LC→SC converted column across swimmers
    stroke_names = list(dict.fromkeys(name for name, _ in EVENTS))
    for stroke in stroke_names:
        conv_key = (stroke, "LC→SC")
        entries = []
        for idx, row in enumerate(rows):
            ev = row["events"].get(conv_key)
            if ev:
                secs = time_to_seconds(ev["time_str"])
                if secs:
                    entries.append((secs, idx))
        entries.sort(key=lambda x: x[0])
        for rank, (_, idx) in enumerate(entries, 1):
            if conv_key in rows[idx]["events"]:
                rows[idx]["events"][conv_key]["club_rank"] = rank

    rows.sort(key=lambda r: (r["avg_rank"], -r["event_count"]))
    return rows


# ── Update PBs in spreadsheet ─────────────────────────────────────────────────

def update_pbs(swimmers):
    """Update 2026-Times tab with faster times from site (Margot only)."""
    import datetime as dt
    margot = next((s for s in swimmers if s["is_margot"]), None)
    if not margot:
        print("  Margot not found in swimmer list — skipping PB update")
        return

    print(f"\nUpdating PBs in 2026-Times...")
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["2026-Times"]
    updated = 0

    for row in ws.iter_rows(min_row=4, max_row=45):
        course_cell = row[1]
        event_cell  = row[2]
        pb_cell     = row[3]
        date_cell   = row[4]

        if not course_cell.value or not event_cell.value:
            continue

        course     = "SC" if str(course_cell.value).strip() == "S" else "LC"
        short_name = str(event_cell.value).strip()
        site_data  = margot["bests"].get((short_name, course))
        if not site_data:
            continue

        sheet_secs = cell_to_seconds(pb_cell.value)
        site_secs  = site_data["time_secs"]

        if sheet_secs is None or site_secs < sheet_secs:
            total   = site_secs
            h       = int(total // 3600)
            m       = int((total % 3600) // 60)
            s       = total % 60
            s_int   = int(s)
            us      = round((s - s_int) * 1_000_000)
            pb_cell.value = dt.time(h, m, s_int, us)
            try:
                date_cell.value = datetime.strptime(site_data["date_str"], "%d/%m/%y")
            except ValueError:
                pass
            old = seconds_to_str(sheet_secs) if sheet_secs else "-"
            print(f"    {course} {short_name}: {old} → {site_data['time_str']}")
            updated += 1

    wb.save(XLSX_PATH)
    print(f"  → {updated} PBs updated")


# ── Excel: Club Rankings tab ──────────────────────────────────────────────────

def build_excel_tab(rows, run_time):
    print("\nBuilding Club Rankings tab...")
    wb = openpyxl.load_workbook(XLSX_PATH)
    if "Club Rankings" in wb.sheetnames:
        del wb["Club Rankings"]
    ws = wb.create_sheet("Club Rankings")

    yellow_fill = PatternFill("solid", fgColor="FFFF99")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    stroke_fill = PatternFill("solid", fgColor="2E75B6")
    hfont       = Font(color="FFFFFF", bold=True, size=10)
    bfont       = Font(bold=True, size=10)
    nfont       = Font(size=10)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left", vertical="center")
    thin        = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Stroke group spans for row 2 (3 columns per stroke: SC, LC, LC→SC)
    stroke_groups = []
    i = 0
    while i < len(EVENTS):
        name = EVENTS[i][0]
        span = sum(1 for e in EVENTS[i:] if e[0] == name)
        stroke_groups.append((name, span))
        i += span

    last_col = 3 + len(EVENTS)

    # Row 1: timestamp
    c = ws.cell(row=1, column=1, value=f"Last updated: {run_time}")
    c.font = Font(italic=True, size=9, color="666666")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    # Row 2: stroke group headers
    col = 3
    for stroke_name, span in stroke_groups:
        cell = ws.cell(row=2, column=col, value=stroke_name)
        cell.fill = stroke_fill; cell.font = Font(color="FFFFFF", bold=True, size=9)
        cell.alignment = center; cell.border = thin
        if span > 1:
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+span-1)
            for c2 in range(col+1, col+span):
                ws.cell(row=2, column=c2).fill = stroke_fill
                ws.cell(row=2, column=c2).border = thin
        col += span
    c = ws.cell(row=2, column=last_col, value="Avg Rank")
    c.fill = header_fill; c.font = hfont; c.alignment = center; c.border = thin

    # Row 3: column headers
    for col_num, label in [(1, "Swimmer"), (2, "Events")]:
        c = ws.cell(row=3, column=col_num, value=label)
        c.fill = header_fill; c.font = hfont
        c.alignment = left if col_num == 1 else center; c.border = thin
    for j, (name, course) in enumerate(EVENTS):
        c = ws.cell(row=3, column=3+j, value=course)
        c.fill = header_fill; c.font = hfont; c.alignment = center; c.border = thin
    c = ws.cell(row=3, column=last_col, value="Avg Rank")
    c.fill = header_fill; c.font = hfont; c.alignment = center; c.border = thin

    # Data rows
    for r_idx, swimmer in enumerate(rows):
        row_num   = 4 + r_idx
        is_margot = swimmer["is_margot"]
        fill      = yellow_fill if is_margot else None
        font      = bfont if is_margot else nfont

        def write(col, value, align=center):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.font = font; cell.alignment = align; cell.border = thin
            if fill:
                cell.fill = fill

        name_display = f"⭐ {swimmer['name']}" if is_margot else swimmer["name"]
        write(1, name_display, left)
        write(2, swimmer["event_count"])

        for j, (short_name, course) in enumerate(EVENTS):
            ev = swimmer["events"].get((short_name, course))
            if ev:
                rank_str = ordinal(ev["club_rank"])
                val = f"{ev['time_str']} ({rank_str})\n{ev['date_str']}"
            else:
                val = ""
            write(3+j, val)

        avg = round(swimmer["avg_rank"], 1) if swimmer["event_count"] > 0 else ""
        write(last_col, avg)

    # Column widths & row heights
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 8
    for j in range(len(EVENTS)):
        ws.column_dimensions[get_column_letter(3+j)].width = 14
    ws.column_dimensions[get_column_letter(last_col)].width = 10
    for r_idx in range(len(rows)):
        ws.row_dimensions[4+r_idx].height = 30
    ws.freeze_panes = "C4"

    wb.save(XLSX_PATH)
    print(f"  → {len(rows)} swimmers, {len(EVENTS)} event columns")


# ── HTML report ───────────────────────────────────────────────────────────────

def slug(name):
    return name.lower().replace(" ", "-").replace("→", "-")


def read_quals(xlsx_path=None, tab_name="2026-Times"):
    """
    Read county/regional QT and qual status from a *-Times tab.
    Returns dict keyed by (short_name, course).
    """
    import datetime as _dt
    wb = openpyxl.load_workbook(xlsx_path or XLSX_PATH, data_only=True)
    ws = wb[tab_name]
    result = {}
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=False):
        course_val = row[1].value
        event_val  = row[2].value
        if not course_val or not event_val:
            continue
        course     = "SC" if str(course_val).strip() == "S" else "LC"
        event      = str(event_val).strip()
        r          = row[0].row

        def _secs(val):
            if val is None: return None
            if isinstance(val, _dt.time):
                return val.hour*3600 + val.minute*60 + val.second + val.microsecond/1e6
            if isinstance(val, _dt.timedelta): return val.total_seconds()
            return None

        pb_secs    = _secs(ws.cell(r, 4).value)
        cty_qt     = _secs(ws.cell(r, 7).value)
        cty_cons   = _secs(ws.cell(r, 9).value)
        reg_qt     = _secs(ws.cell(r, 13).value)
        reg_cons   = _secs(ws.cell(r, 15).value)

        # Compute Q/C directly — formula cache is empty when written by openpyxl
        def qual_status(pb, qt, cons):
            if pb and qt and pb <= qt:   return "Q"
            if pb and cons and pb <= cons: return "C"
            return ""

        # County qual checks both SC and LC — look up the paired course
        paired = "LC" if course == "SC" else "SC"
        paired_key = (event, paired)

        result[(event, course)] = {
            "county_qt":   cty_qt,
            "county_cons": cty_cons,
            "region_qt":   reg_qt,
            "region_cons": reg_cons,
            "pb_secs":     pb_secs,
            "_county_qual_pending": qual_status(pb_secs, cty_qt, cty_cons),
            "_region_qual_pending": qual_status(pb_secs, reg_qt, reg_cons),
        }

    # Second pass: combine SC+LC for Q/C (qualifies if either course meets the time)
    for (event, course), d in result.items():
        paired = "LC" if course == "SC" else "SC"
        paired_d = result.get((event, paired), {})
        def combined(key):
            a = d.get(key, "")
            b = paired_d.get(key, "")
            if "Q" in (a, b): return "Q"
            if "C" in (a, b): return "C"
            return ""
        d["county_qual"] = combined("_county_qual_pending")
        d["region_qual"] = combined("_region_qual_pending")

    return result

def read_margot_quals():
    """Alias — reads age-13 targets from 2027-Times for forward planning."""
    return read_quals(XLSX_PATH, "2027-Times")

def read_ava_quals():
    """Read youngest-group QTs from Ava's spreadsheet."""
    return read_quals(AVA_XLSX_PATH, "2026-Times")


def read_age_group_quals(year_of_birth):
    """
    Read county/regional qualifying times for a given birth year.
    Uses 2027 as the championship year (next season — same basis as Brompton app).
    Returns dict keyed by (event, course), or empty dict if age is out of range.
    """
    import datetime as _dt

    CHAMP_YEAR = 2027
    age = CHAMP_YEAR - year_of_birth

    # County age bands (Middlesex County Championships)
    # Sheet cols: "10+11"(gi=0), "12"(gi=1), "13"(gi=2), "14"(gi=3), "15"(gi=4), "16"(gi=5), "17+"(gi=6)
    COUNTY_BANDS = ["10+11", "12", "13", "14", "15", "16", "17+"]
    if age >= 17:          county_band = "17+"
    elif age == 16:        county_band = "16"
    elif age == 15:        county_band = "15"
    elif age == 14:        county_band = "14"
    elif age == 13:        county_band = "13"
    elif age == 12:        county_band = "12"
    elif age in (10, 11):  county_band = "10+11"
    else:                  county_band = None  # too young for county

    # Regional age bands (SE London Summer Championships — LC only)
    # Sheet row 3 labels: "11-12", "13", "14", "15", "16", "17", "18+"
    if age >= 18:          reg_band = "18+"
    elif age == 17:        reg_band = "17"
    elif age == 16:        reg_band = "16"
    elif age == 15:        reg_band = "15"
    elif age == 14:        reg_band = "14"
    elif age == 13:        reg_band = "13"
    elif age in (11, 12):  reg_band = "11-12"
    else:                  reg_band = None  # too young for regional

    def _secs(val):
        if val is None: return None
        if isinstance(val, _dt.time):
            return val.hour*3600 + val.minute*60 + val.second + val.microsecond/1e6
        if isinstance(val, _dt.timedelta): return val.total_seconds()
        return None

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    # ── County Times ─────────────────────────────────────────────────────────
    county_data = {}
    if county_band:
        gi         = COUNTY_BANDS.index(county_band)
        lc_qt_col  = 2 + gi * 2   # 1-indexed openpyxl column
        lc_con_col = 3 + gi * 2
        sc_qt_col  = 16 + gi * 2
        sc_con_col = 17 + gi * 2
        cws = wb["County Times 2026"]
        for row in cws.iter_rows(min_row=5, max_row=cws.max_row, values_only=False):
            ev_name = row[0].value
            if not ev_name: continue
            r = row[0].row
            county_data[str(ev_name).strip()] = {
                "LC": (_secs(cws.cell(r, lc_qt_col).value),  _secs(cws.cell(r, lc_con_col).value)),
                "SC": (_secs(cws.cell(r, sc_qt_col).value),  _secs(cws.cell(r, sc_con_col).value)),
            }

    # ── Regional Times ────────────────────────────────────────────────────────
    # Sheet structure: row 3 = age band labels, row 4 = QT/Con, rows 5+ = events in col A.
    # Regional championships are LC only — no SC columns.
    reg_data = {}
    if reg_band:
        rws      = wb["Regional Times 2026"]
        band_row = [c.value for c in list(rws.iter_rows(min_row=3, max_row=3))[0]]
        # Find the QT column for this age band (band label is at the QT col; Con is at QT+1)
        lc_qt_col = lc_con_col = None
        for ci, val in enumerate(band_row):
            if str(val or "").strip() == reg_band:
                lc_qt_col  = ci + 1   # openpyxl 1-indexed
                lc_con_col = ci + 2
                break
        if lc_qt_col:
            for row in rws.iter_rows(min_row=5, max_row=rws.max_row, values_only=False):
                ev_name = row[0].value
                if not ev_name: continue
                r = row[0].row
                reg_data[str(ev_name).strip()] = {
                    "LC": (_secs(rws.cell(r, lc_qt_col).value), _secs(rws.cell(r, lc_con_col).value)),
                    "SC": (None, None),
                }

    # ── Combine ───────────────────────────────────────────────────────────────
    result = {}
    for event in set(county_data) | set(reg_data):
        for course in ("SC", "LC"):
            cty = county_data.get(event, {}).get(course, (None, None))
            reg = reg_data.get(event, {}).get(course, (None, None))
            result[(event, course)] = {
                "county_qt":   cty[0], "county_cons": cty[1],
                "region_qt":   reg[0], "region_cons": reg[1],
            }
    return result


def fetch_peer_swimmers(config_path, year_of_birth):
    """Fetch PBs for a peer age group."""
    config = json.loads(config_path.read_text())
    swimmers = config["swimmers"]
    results = []
    total = len(swimmers)
    for i, s in enumerate(swimmers, 1):
        print(f"  [{i}/{total}] {s['name']}...")
        bests = fetch_swimmer_bests(s["member_id"], s["name"])
        print(f"    → {len(bests)} times")
        results.append({
            "name":          s["name"],
            "member_id":     s["member_id"],
            "is_margot":     False,
            "year_of_birth": year_of_birth,
            "has_history":   True,
            "bests":         bests,
        })
        if i < total:
            time.sleep(RATE_DELAY)
    return results


def build_peer_rows(peer_swimmers):
    """Build unranked rows for a peer age group."""
    rows = []
    for sw in peer_swimmers:
        events_out = {}
        rank_count = 0
        for key, data in sw["bests"].items():
            events_out[key] = {
                "time_str":  data["time_str"],
                "date_str":  data["date_str"],
                "club_rank": None,
            }
            rank_count += 1
            if key[1] == "LC":
                converted = data.get("converted", "")
                conv_key  = (key[0], "LC→SC")
                if converted and converted != "-":
                    events_out[conv_key] = {
                        "time_str":  converted,
                        "date_str":  data["date_str"],
                        "club_rank": None,
                    }
        rows.append({
            "name":          sw["name"],
            "is_margot":     False,
            "year_of_birth": sw["year_of_birth"],
            "has_history":   True,
            "avg_rank":      999,
            "event_count":   rank_count,
            "events":        events_out,
        })
    rows.sort(key=lambda r: (-r["event_count"], r["name"]))
    return rows


PREV_BESTS_PATH = MARGOT_DIR / "margot_previous_bests.json"

def load_previous_bests():
    import json as _j
    if PREV_BESTS_PATH.exists():
        return _j.loads(PREV_BESTS_PATH.read_text())
    return {}

def save_current_bests(swimmers):
    import json as _j
    margot = next((s for s in swimmers if s["is_margot"]), None)
    if not margot: return
    snapshot = {
        f"{ev}|{course}": data["time_str"]
        for (ev, course), data in margot["bests"].items()
    }
    PREV_BESTS_PATH.write_text(_j.dumps(snapshot, indent=2))


def fetch_peer_histories(swimmers, cache_path, highlight_name):
    """
    Fetch full time history for every swimmer in the list.
    Highlight swimmer (Margot/Ava) is fetched first so their data is always fresh.
    Saves to cache_path as JSON. Returns dict {name: {event|course: [swims]}}.
    """
    import json as _j

    # Put highlight swimmer first so they're always updated
    ordered = sorted(swimmers, key=lambda s: 0 if s["name"] == highlight_name else 1)

    all_hist = {}
    total = len(ordered)
    for i, swimmer in enumerate(ordered, 1):
        name = swimmer["name"]
        print(f"  [{i}/{total}] {name} history...")
        hist = {}
        seen = set()
        for short_name, course in [(n, c) for n, c in EVENTS if c != "LC→SC"]:
            if (short_name, course) in seen:
                continue
            seen.add((short_name, course))
            # Skip if swimmer has no PB for this event (saves requests)
            if not swimmer["bests"].get((short_name, course)):
                continue
            stroke_code = STROKE_CODES.get(short_name)
            if not stroke_code:
                continue
            pool = "S" if course == "SC" else "L"
            url = (
                f"https://www.swimmingresults.org/individualbest/personal_best_time_date.php"
                f"?tiref={swimmer['member_id']}&mode=A&tstroke={stroke_code}&tcourse={pool}"
                f"&Pool={pool}&Stroke={stroke_code}"
            )
            soup = curl_get(url, referer="https://www.swimmingresults.org/individualbest/")
            table = soup.find("table", id="rankTable")
            rows = []
            if table:
                for row in table.find_all("tr")[1:]:
                    cells = [td.get_text(strip=True) for td in row.find_all("td")]
                    if len(cells) >= 4:
                        rows.append({
                            "time": cells[0], "pts": cells[1],
                            "round": cells[2], "date": cells[3],
                            "meet": cells[4] if len(cells) > 4 else "",
                        })
            hist[f"{short_name}|{course}"] = rows
        all_hist[name] = hist
        print(f"    → {sum(len(v) for v in hist.values())} swims")

    cache_path.write_text(_j.dumps(all_hist, indent=2))
    print(f"  Saved to {cache_path.name}")
    return all_hist


def load_peer_histories(cache_path):
    """Load cached peer histories. Returns empty dict if cache missing."""
    import json as _j
    if cache_path.exists():
        return _j.loads(cache_path.read_text())
    return {}


def fetch_ava_club_swimmers():
    """Fetch PBs for all swimmers in swimmers_config_ava.json."""
    config = json.loads(AVA_CONFIG_PATH.read_text())
    swimmers = config["swimmers"]
    results = []
    total = len(swimmers)
    for i, s in enumerate(swimmers, 1):
        print(f"  [{i}/{total}] {s['name']}...")
        bests = fetch_swimmer_bests(s["member_id"], s["name"])
        print(f"    → {len(bests)} times")
        results.append({
            "name":          s["name"],
            "member_id":     s["member_id"],
            "is_margot":     s["name"] == AVA_NAME,   # highlights Ava in the HTML
            "year_of_birth": config.get("year_of_birth", 2017),
            "bests":         bests,
        })
        if i < total:
            time.sleep(RATE_DELAY)
    return results


def fetch_ava_history():
    """Fetch all recorded swims for Ava across all events."""
    AVA_ID = "1846170"
    history = {}
    seen = set()
    for short_name, course in [(n, c) for n, c in EVENTS if c != "LC→SC"]:
        if (short_name, course) in seen:
            continue
        seen.add((short_name, course))
        stroke_code = STROKE_CODES.get(short_name)
        if not stroke_code:
            continue
        pool = "S" if course == "SC" else "L"
        url = (
            f"https://www.swimmingresults.org/individualbest/personal_best_time_date.php"
            f"?tiref={AVA_ID}&mode=A&tstroke={stroke_code}&tcourse={pool}"
            f"&Pool={pool}&Stroke={stroke_code}"
        )
        soup = curl_get(url, referer="https://www.swimmingresults.org/individualbest/")
        table = soup.find("table", id="rankTable")
        rows = []
        if table:
            for row in table.find_all("tr")[1:]:
                cells = [td.get_text(strip=True) for td in row.find_all("td")]
                if len(cells) >= 4:
                    rows.append({
                        "time":  cells[0], "pts": cells[1],
                        "round": cells[2], "date": cells[3],
                        "meet":  cells[4] if len(cells) > 4 else "",
                    })
        history[(short_name, course)] = rows
    return history


def build_ava_xlsx(ava_swimmers):
    """
    Update AvaSwimTimes.xlsx with Ava's latest PBs.
    The file is a copy of MargotSwimTimes.xlsx — same tabs, same format.
    Only updates PB (col D) and date (col E) in 2026-Times, where site has a faster time.
    """
    import shutil as _sh
    import datetime as _dt

    # Bootstrap: if file doesn't exist or is missing tabs, rebuild from Margot's template
    if not AVA_XLSX_PATH.exists() or len(openpyxl.load_workbook(AVA_XLSX_PATH).sheetnames) < 4:
        _sh.copy2(XLSX_PATH, AVA_XLSX_PATH)
        wb = openpyxl.load_workbook(AVA_XLSX_PATH)
        YEAR_TABS = [s for s in wb.sheetnames if "Times" in s
                     and s not in ("County Times 2026", "Regional Times 2026")]
        for tab in YEAR_TABS:
            ws = wb[tab]
            for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
                if not str(row[1].value or "").strip(): continue
                r = row[0].row
                for col in [4,5,6,7,9,11,12,13,14,15,16,17,18]:
                    ws.cell(r, col).value = None
                for col in range(19, ws.max_column + 1):
                    ws.cell(r, col).value = None
        wb.save(AVA_XLSX_PATH)

    # Get Ava's bests from the swimmer data
    ava = next((s for s in ava_swimmers if s["is_margot"]), None)
    if not ava:
        print("  Ava not found in swimmer list")
        return

    wb  = openpyxl.load_workbook(AVA_XLSX_PATH)
    ws  = wb["2026-Times"]
    updated = 0

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        course_v = str(row[1].value or "").strip()
        event_v  = str(row[2].value or "").strip()
        if not course_v or not event_v: continue
        course   = "SC" if course_v == "S" else "LC"
        best     = ava["bests"].get((event_v, course))
        if not best: continue
        r        = row[0].row
        curr     = cell_to_seconds(ws.cell(r, 4).value)
        new_secs = best["time_secs"]
        if curr is None or new_secs < curr:
            t = new_secs; m=int(t//60); s=t%60; si=int(s); us=round((s-si)*1_000_000)
            ws.cell(r, 4).value = _dt.time(0,m,si,us)
            ws.cell(r, 4).number_format = "m:ss.00"
            try:
                ws.cell(r, 5).value = datetime.strptime(best["date_str"], "%d/%m/%y")
            except: pass
            updated += 1

    wb.save(AVA_XLSX_PATH)
    print(f"  → AvaSwimTimes.xlsx: {updated} PBs updated")


def build_ava_html(ava_history, run_time):
    """Build ava_times.html — Ava Mandikos personal best progression."""
    import json as _j
    AVA_PATH = AVA_HTML_PATH

    # Collect Ava's best time per event from history
    events_with_data = []
    for short_name, course in [(n, c) for n, c in EVENTS if c != "LC→SC"]:
        swims = ava_history.get((short_name, course), [])
        if swims:
            best = min(swims, key=lambda s: time_to_seconds(s["time"]) or 9999)
            events_with_data.append((short_name, course, best, len(swims)))

    hist_json = _j.dumps({
        f"{n}|{c}": v for (n, c), v in ava_history.items()
    })

    rows_html = ""
    for short_name, course, best, count in events_with_data:
        rows_html += (
            f'<tr onclick="showHistory(\'Ava Mandikos\',\'{short_name}\',\'{course}\')" style="cursor:pointer">'
            f'<td style="padding:6px 12px;font-weight:500">{short_name}</td>'
            f'<td style="padding:6px 12px;text-align:center">{course}</td>'
            f'<td style="padding:6px 12px;text-align:center;font-weight:600">{best["time"]}</td>'
            f'<td style="padding:6px 12px;text-align:center;color:#888">{best["date"]}</td>'
            f'<td style="padding:6px 12px;text-align:center;color:#888">{count}</td>'
            f'</tr>'
        )

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Ava Mandikos — Swimming Times</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/chartjs-adapter-date-fns@3.0.0/dist/chartjs-adapter-date-fns.bundle.min.js"></script>
<style>
  body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;font-size:13px;background:#f5f5f5;margin:0;padding:20px;color:#333}}
  h1{{font-size:20px;font-weight:600;color:#1a3a5c;margin:0 0 4px}}
  .meta{{font-size:11px;color:#888;margin-bottom:20px}}
  table{{border-collapse:collapse;background:white;box-shadow:0 1px 4px rgba(0,0,0,.1);width:100%;max-width:600px}}
  th{{background:#1f4e79;color:white;padding:8px 12px;text-align:left;font-size:11px}}
  tr:nth-child(even){{background:#f9f9f9}}
  tr:hover{{background:#eef4fb}}
</style>
</head>
<body>
<h1>Ava Mandikos — Personal Bests</h1>
<p class="meta">Last updated: {run_time} &nbsp;|&nbsp; Chelsea &amp; Westminster SC &nbsp;|&nbsp; Click any row to see full history</p>
<table>
<thead><tr>
  <th>Event</th><th>Course</th><th>Best Time</th><th>Date</th><th>Swims</th>
</tr></thead>
<tbody>{rows_html}</tbody>
</table>

<div id="history-modal" style="display:none;position:fixed;top:0;left:0;width:100%;height:100%;
  background:rgba(0,0,0,.5);z-index:1000;align-items:center;justify-content:center;">
  <div style="background:white;border-radius:8px;padding:24px;max-width:820px;width:95%;
    max-height:90vh;overflow-y:auto;box-shadow:0 8px 32px rgba(0,0,0,.3)">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:16px">
      <h2 id="modal-title" style="margin:0;font-size:15px;color:#1a3a5c"></h2>
      <button onclick="closeModal()" style="border:none;background:none;font-size:22px;cursor:pointer;color:#666">&times;</button>
    </div>
    <div style="margin-bottom:20px"><canvas id="history-chart" height="120"></canvas></div>
    <table id="modal-table" style="width:100%;border-collapse:collapse;font-size:12px"></table>
  </div>
</div>
{_modal_script(hist_json, "{{}}")}
</body></html>"""
    AVA_PATH.write_text(html, encoding="utf-8")
    print(f"  → ava_times.html written ({len(events_with_data)} events)")


def _modal_script(history_json, qt_json):
    """Return the shared modal+chart JS block."""
    return f"""<script>
var HISTORY={history_json};
var QT_DATA={qt_json};
var _chart=null;

function parseSeconds(t){{
  if(!t)return null; t=t.trim();
  return t.indexOf(':')>=0 ? parseFloat(t.split(':')[0])*60+parseFloat(t.split(':')[1]) : parseFloat(t);
}}
function formatTime(s){{
  if(s===null||isNaN(s))return '';
  s=Math.round(s*100)/100;
  if(s>=60){{var m=Math.floor(s/60);var r=(s-m*60).toFixed(2).padStart(5,'0');return m+':'+r;}}
  return s.toFixed(2);
}}
function parseDate(d){{var p=d.split('/');return new Date('20'+p[2],p[1]-1,p[0]);}}

function showHistory(swimmerName,eventName,course){{
  var swimmerHist=HISTORY[swimmerName]||{{}};
  var scSwims=swimmerHist[eventName+'|SC']||[];
  var lcSwims=swimmerHist[eventName+'|LC']||[];
  var swims  =course==='SC'?scSwims:lcSwims;
  document.getElementById('modal-title').textContent=swimmerName+' — '+eventName+' '+course+' ('+swims.length+' swims)';

  // No cached history for this swimmer yet
  if(!HISTORY[swimmerName]&&swimmerName!==''){{
    document.getElementById('history-modal').style.display='flex';
    document.getElementById('modal-table').innerHTML=
      '<tr><td style="padding:20px;text-align:center;color:#888;font-style:italic">'
      +'No history cached for '+swimmerName+'.<br>Run <strong>python3 swimming_results.py --fetch-history</strong> to load histories for all BSC swimmers.'
      +'</td></tr>';
    if(_chart){{_chart.destroy();_chart=null;}}
    document.getElementById('chart-legend').innerHTML='';
    return;
  }}

  // Sort chronologically
  function sortedArr(arr){{
    return arr.slice().sort(function(a,b){{return parseDate(a.date)-parseDate(b.date);}});
  }}
  var scArr=sortedArr(scSwims), lcArr=sortedArr(lcSwims);

  // Each dataset uses x: ISO date, y: seconds — so SC and LC sit at their own correct dates
  function toPoints(arr){{
    return arr.map(function(s){{return{{x:parseDate(s.date).toISOString(),y:parseSeconds(s.time)}};}});
  }}
  var scPoints=toPoints(scArr), lcPoints=toPoints(lcArr);
  // For y-axis scale calculation only
  var scVals=scArr.map(function(s){{return parseSeconds(s.time);}});
  var lcVals=lcArr.map(function(s){{return parseSeconds(s.time);}});

  var qtKey=eventName+'|'+course;
  var qd=QT_DATA[qtKey]||{{}};
  var ctyQt=qd.county_qt||null, ctyCons=qd.county_cons||null;
  var regQt=qd.region_qt||null, regCons=qd.region_cons||null;

  // Y-axis scale from all data + all 4 reference lines
  var allVals=scVals.concat(lcVals).filter(function(v){{return v!==null&&!isNaN(v);}});
  [ctyQt,ctyCons,regQt,regCons].forEach(function(v){{if(v)allVals.push(v);}});
  var yMin=Math.min.apply(null,allVals),yMax=Math.max.apply(null,allVals)||1;
  var range=yMax-yMin||1,rawStep=range/5;
  var mag=Math.pow(10,Math.floor(Math.log10(rawStep)));
  var step=[0.5,1,2,2.5,5,10,15,20,30,60].find(function(s){{return s*mag>=rawStep;}})*mag||rawStep;
  var axMin=Math.floor(yMin/step)*step, axMax=Math.ceil(yMax/step)*step+step;

  // Build datasets — both SC and LC use date-based x so they sit at their own correct positions
  var allDates=[].concat(
    scArr.map(function(s){{return parseDate(s.date);}}).concat(lcArr.map(function(s){{return parseDate(s.date);}}))
  );
  var minD=allDates.length?new Date(Math.min.apply(null,allDates)-30*864e5):new Date(2025,0,1);
  var maxD=allDates.length?new Date(Math.max.apply(null,allDates)+30*864e5):new Date(2027,0,1);
  function refLinePoints(val){{return [{{x:minD.toISOString(),y:val}},{{x:maxD.toISOString(),y:val}}];}}

  var datasets=[];
  var swimPoints=course==='SC'?scPoints:lcPoints;
  var swimColor=course==='SC'?'#2e75b6':'#27ae60';
  var swimBg=course==='SC'?'rgba(46,117,182,0.08)':'rgba(39,174,96,0.08)';
  if(swimPoints.length) datasets.push({{label:course,data:swimPoints,borderColor:swimColor,backgroundColor:swimBg,pointRadius:5,pointHoverRadius:7,tension:0.3,fill:false}});
  if(ctyCons) datasets.push({{label:'County Cons',data:refLinePoints(ctyCons),borderColor:'#e67e22',borderDash:[4,4],pointRadius:0,borderWidth:1.5}});
  if(ctyQt)   datasets.push({{label:'County QT',  data:refLinePoints(ctyQt), borderColor:'#e74c3c',borderDash:[6,3],pointRadius:0,borderWidth:1.5}});
  if(regCons) datasets.push({{label:'Regional Cons',data:refLinePoints(regCons),borderColor:'#16a085',borderDash:[4,4],pointRadius:0,borderWidth:1.5}});
  if(regQt)   datasets.push({{label:'Regional QT', data:refLinePoints(regQt), borderColor:'#8e44ad',borderDash:[6,3],pointRadius:0,borderWidth:1.5}});

  if(_chart){{_chart.destroy();_chart=null;}}
  var ctx=document.getElementById('history-chart').getContext('2d');

  _chart=new Chart(ctx,{{
    type:'line',
    data:{{datasets:datasets}},
    options:{{
      responsive:true,
      maintainAspectRatio:false,
      plugins:{{
        legend:{{display:false}},
        tooltip:{{callbacks:{{label:function(c){{return c.dataset.label+': '+formatTime(c.parsed.y);}}}}}}
      }},
      scales:{{
        y:{{reverse:true,min:axMin,max:axMax,
            ticks:{{stepSize:step,callback:function(v){{return formatTime(v);}},font:{{size:11}}}},
            title:{{display:true,text:'Time (faster ↑)',font:{{size:11}}}}}},
        x:{{type:'time',time:{{unit:'month',displayFormats:{{month:'MMM yy'}}}},
            ticks:{{font:{{size:10}},maxRotation:45}}}}
      }}
    }}
  }});

  // Custom checkbox legend
  var legendDiv=document.getElementById('chart-legend');
  legendDiv.innerHTML='';
  _chart.data.datasets.forEach(function(ds,i){{
    var colMap={{'County Cons':[4,5],'County QT':[6,7],'Regional Cons':[8,9],'Regional QT':[10,11]}};
    if(!colMap[ds.label]) return; // skip swim data line — no checkbox needed
    var label=document.createElement('label');
    label.style.cssText='display:flex;align-items:center;gap:5px;cursor:pointer;user-select:none';
    var cb=document.createElement('input');
    cb.type='checkbox'; cb.checked=true;
    // Table col indices (0-based from first th): 0=Date,1=Meet,2=Time,3=vLead,
    //   4=CtyCons,5=CtyConsGap, 6=CtyQT,7=CtyQTGap, 8=RegCons,9=RegConsGap, 10=RegQT,11=RegQTGap
    cb.addEventListener('change',function(){{
      _chart.setDatasetVisibility(i,cb.checked);
      _chart.update();
      svg.style.opacity=cb.checked?'1':'0.35';
      var cols=colMap[ds.label];
      if(cols){{
        var table=document.getElementById('modal-table');
        if(table){{
          var allRows=table.querySelectorAll('tr');
          allRows.forEach(function(row){{
            cols.forEach(function(ci){{
              if(row.cells[ci]) row.cells[ci].style.display=cb.checked?'':'none';
            }});
          }});
        }}
      }}
    }});
    var svgNS='http://www.w3.org/2000/svg';
    var svg=document.createElementNS(svgNS,'svg');
    svg.setAttribute('width','28'); svg.setAttribute('height','10');
    var line=document.createElementNS(svgNS,'line');
    line.setAttribute('x1','0'); line.setAttribute('y1','5');
    line.setAttribute('x2','28'); line.setAttribute('y2','5');
    line.setAttribute('stroke',ds.borderColor);
    line.setAttribute('stroke-width',ds.borderWidth||2);
    if(ds.borderDash&&ds.borderDash.length) line.setAttribute('stroke-dasharray',ds.borderDash.join(','));
    svg.appendChild(line);
    var txt=document.createTextNode(ds.label);
    label.appendChild(cb); label.appendChild(svg); label.appendChild(txt);
    legendDiv.appendChild(label);
  }});

  // Table — most recent first
  var sorted=swims.slice().sort(function(a,b){{return parseDate(b.date)-parseDate(a.date);}});
  var ctyConsFmt=formatTime(ctyCons)||'—', ctyQtFmt=formatTime(ctyQt)||'—';
  var regConsFmt=formatTime(regCons)||'—', regQtFmt=formatTime(regQt)||'—';

  function gapStr(swimSecs,refSecs){{
    if(!refSecs||!swimSecs) return '—';
    var diff=Math.round((swimSecs-refSecs)*100)/100;
    if(Math.abs(diff)<0.005) return '✓';
    var sign=diff>0?'+':'-';
    var abs=Math.abs(diff);
    if(abs>=60){{var m=Math.floor(abs/60);var s=(abs-m*60).toFixed(1);return sign+m+':'+s.padStart(4,'0');}}
    return sign+abs.toFixed(2)+'s';
  }}

  // Age-group club leader time from pre-computed rankings
  var leaderSecs=qd.leader_secs||null;
  var leaderName=qd.leader_name||null;

  // Find PB
  var pbSecs=Math.min.apply(null,sorted.map(function(s){{return parseSeconds(s.time)||Infinity;}}).filter(isFinite));

  var tbody='<thead><tr style="background:#1f4e79;color:white">'
    +'<th style="padding:6px 8px;text-align:left">Date</th>'
    +'<th style="padding:6px 8px;text-align:left">Meet</th>'
    +'<th style="padding:6px 8px;text-align:center">Time</th>'
    +'<th style="padding:6px 8px;text-align:center" title="'+(leaderName?'Gap to '+leaderName+' (age-group leader)':'No leader data')+'">v Lead</th>'
    +'<th style="padding:6px 8px;text-align:center;color:#ffa94d">Cty Cons</th>'
    +'<th style="padding:6px 8px;text-align:center;color:#ffa94d">Gap</th>'
    +'<th style="padding:6px 8px;text-align:center;color:#ff6b6b">Cty QT</th>'
    +'<th style="padding:6px 8px;text-align:center;color:#ff6b6b">Gap</th>'
    +'<th style="padding:6px 8px;text-align:center;color:#63e6be">Reg Cons</th>'
    +'<th style="padding:6px 8px;text-align:center;color:#63e6be">Gap</th>'
    +'<th style="padding:6px 8px;text-align:center;color:#cc99ff">Reg QT</th>'
    +'<th style="padding:6px 8px;text-align:center;color:#cc99ff">Gap</th>'
    +'</tr></thead><tbody>';

  sorted.forEach(function(s,i){{
    var secs=parseSeconds(s.time);
    var isPB=secs&&Math.abs(secs-pbSecs)<0.005;
    var bg=isPB?'#fffde7':(i%2===0?'#f9f9f9':'white');
    function qualTd(fmt,ref){{
      var col=ref&&secs&&secs<=ref?'color:#27ae60;font-weight:600':'color:#888';
      return '<td style="padding:5px 8px;text-align:center;'+col+'">'+fmt+'</td>';
    }}
    function gapTd(ref){{
      var g=gapStr(secs,ref);
      var col=g==='✓'?'color:#27ae60;font-weight:600':g.startsWith('-')?'color:#27ae60':g==='—'?'color:#ccc':'color:#e74c3c';
      return '<td style="padding:4px 8px;text-align:center;font-size:10px;'+col+'">'+g+'</td>';
    }}
    var vLead=gapStr(secs,leaderSecs);
    var vLeadCol=vLead==='—'||vLead==='✓'?'color:#888':vLead.startsWith('+')?'color:#e74c3c':'color:#27ae60';
    tbody+='<tr data-ts="'+parseDate(s.date).getTime()+'" style="background:'+bg+'">'
      +'<td style="padding:5px 8px">'+s.date+'</td>'
      +'<td style="padding:5px 8px;max-width:180px;overflow:hidden;text-overflow:ellipsis">'+s.meet+'</td>'
      +'<td style="padding:5px 8px;text-align:center;font-weight:600">'+(isPB?'★ ':'')+s.time+'</td>'
      +'<td style="padding:4px 8px;text-align:center;font-size:10px;'+vLeadCol+'">'+vLead+'</td>'
      +qualTd(ctyConsFmt,ctyCons)+gapTd(ctyCons)
      +qualTd(ctyQtFmt,ctyQt)+gapTd(ctyQt)
      +qualTd(regConsFmt,regCons)+gapTd(regCons)
      +qualTd(regQtFmt,regQt)+gapTd(regQt)
      +'</tr>';
  }});
  document.getElementById('modal-table').innerHTML=tbody+'</tbody>';

  // Qual gap summary — two compact rows (one per course)
  var _qsEl=document.getElementById('modal-qual-summary');
  if(_qsEl){{
    function _bestSwim(arr){{
      if(!arr.length)return null;
      return arr.reduce(function(b,s){{return(!b||parseSeconds(s.time)<parseSeconds(b.time))?s:b;}});
    }}
    var _scBest=_bestSwim(scArr),_lcBest=_bestSwim(lcArr);
    var _scPb=_scBest?parseSeconds(_scBest.time):null;
    var _lcPb=_lcBest?parseSeconds(_lcBest.time):null;
    var _qdSC=QT_DATA[eventName+'|SC']||{{}};
    var _qdLC=QT_DATA[eventName+'|LC']||{{}};

    function _chip(lbl,std,pb){{
      if(!std)return '';
      var g=gapStr(pb,std);
      var ok=g==='✓'||(g!=='—'&&g.charAt(0)==='-');
      var noPb=!pb;
      var gc=noPb?'#bbb':ok?'#15803d':'#dc2626';
      var bg=ok&&!noPb?'rgba(21,128,61,0.08)':'transparent';
      return '<span style="display:inline-flex;align-items:center;gap:3px;padding:1px 6px;border-radius:4px;background:'+bg+'">'
        +'<span style="font-size:10px;color:#888">'+lbl+'</span>'
        +'<span style="font-size:11px;font-weight:700;color:'+gc+';font-variant-numeric:tabular-nums">'+g+'</span>'
        +'</span>';
    }}

    function _row(icon,label,accentBg,accentText,qd,best,pb){{
      var pbPart=best
        ?'<span style="font-size:11px;font-weight:700;color:#1a3a5c;font-variant-numeric:tabular-nums;margin-right:2px">'+best.time+'</span>'
         +'<span style="font-size:10px;color:#aaa;margin-right:8px">'+best.date+'</span>'
        :'<span style="font-size:10px;color:#ccc;font-style:italic;margin-right:8px">no swims</span>';
      var chips=_chip('Cty C',qd.county_cons||null,pb)
        +_chip('Cty QT',qd.county_qt||null,pb)
        +'<span style="color:#ddd;margin:0 2px">│</span>'
        +_chip('Reg C',qd.region_cons||null,pb)
        +_chip('Reg QT',qd.region_qt||null,pb);
      return '<div style="display:flex;align-items:center;gap:6px;padding:4px 8px;border-radius:6px;'
        +'background:'+accentBg+';border:1px solid '+accentText+'20;flex-wrap:wrap">'
        +'<span style="font-size:10px;font-weight:700;color:'+accentText+';white-space:nowrap;min-width:70px">'+icon+' '+label+'</span>'
        +'<span style="color:#e0e0e0;font-size:11px">│</span>'
        +pbPart
        +chips
        +'</div>';
    }}

    _qsEl.style.display='';
    _qsEl.innerHTML=
       _row('📏','Short Course','#eef4fb','#2e75b6',_qdSC,_scBest,_scPb)
      +'<div style="height:4px"></div>'
      +_row('🌊','Long Course','#edfaf3','#1e8449',_qdLC,_lcBest,_lcPb);
  }}

  document.getElementById('history-modal').style.display='flex';
}}

function closeModal(){{
  document.getElementById('history-modal').style.display='none';
  if(_chart){{_chart.destroy();_chart=null;}}
}}
function filterChartPeriod(months,btn){{
  document.querySelectorAll('#chart-period-btns button').forEach(function(b){{
    b.style.background=b===btn?'#1f4e79':'#f0f4f8';
    b.style.color=b===btn?'white':'#444';
    b.style.borderColor=b===btn?'#1f4e79':'#ccc';
  }});
  if(!_chart)return;
  var now=new Date();
  var cutTs=months?new Date(now.getFullYear(),now.getMonth()-months,now.getDate()).getTime():0;
  if(months===null){{
    _chart.options.scales.x.min=undefined;
    _chart.options.scales.x.max=undefined;
    _chart.options.scales.y.min=undefined;
    _chart.options.scales.y.max=undefined;
  }}else{{
    var cut=new Date(now.getFullYear(),now.getMonth()-months,now.getDate());
    _chart.options.scales.x.min=cut.getTime();
    _chart.options.scales.x.max=now.getTime();
    // Recalculate Y range from visible data points only
    var visVals=[];
    _chart.data.datasets.forEach(function(ds){{
      if(!_chart.isDatasetVisible(_chart.data.datasets.indexOf(ds)))return;
      ds.data.forEach(function(pt){{
        var t=new Date(pt.x).getTime();
        if(t>=cutTs&&pt.y!=null)visVals.push(pt.y);
      }});
    }});
    if(visVals.length){{var yMin=Math.min.apply(null,visVals),yMax=Math.max.apply(null,visVals);
      var pad=(yMax-yMin)||2;pad=Math.max(pad*0.15,0.5);
      _chart.options.scales.y.min=Math.floor((yMin-pad)*10)/10;
      _chart.options.scales.y.max=Math.ceil((yMax+pad)*10)/10;
    }}
  }}
  _chart.update();
  document.querySelectorAll('#modal-table tbody tr').forEach(function(row){{
    var ts=parseInt(row.dataset.ts,10);
    row.style.display=(months!==null&&!isNaN(ts)&&ts<cutTs)?'none':'';
  }});
}}

function toggleFullscreen(){{
  var el=document.getElementById('modal-inner');
  var isMax=el.dataset.max==='1';
  if(isMax){{
    el.style.width='92vw'; el.style.height='88vh';
    el.style.borderRadius='8px'; el.dataset.max='0';
  }} else {{
    el.style.width='100vw'; el.style.height='100vh';
    el.style.borderRadius='0'; el.dataset.max='1';
  }}
  if(_chart) _chart.resize();
}}
var _modalDragStartedOnChild=false;
['mousedown','touchstart','pointerdown'].forEach(function(ev){{document.getElementById('history-modal').addEventListener(ev,function(e){{_modalDragStartedOnChild=(e.target!==this);}});}});
document.getElementById('history-modal').addEventListener('click',function(e){{if(e.target===this&&!_modalDragStartedOnChild)closeModal();_modalDragStartedOnChild=false;}});
document.addEventListener('keydown',function(e){{if(e.key==='Escape')closeModal();}});
</script>"""


def _compute_swimmer_quals(rows, default_quals, age_group_quals=None):
    """
    For every swimmer compute combined county_qual and region_qual per (event, course).
    age_group_quals: optional {year_of_birth: quals_dict} — overrides default_quals per age group.
    Returns {swimmer_name: {(event, course): {"county_qual": ..., "region_qual": ...}}}
    """
    def qs(pb, qt, cons):
        if pb and qt and pb <= qt:     return "Q"
        if pb and cons and pb <= cons: return "C"
        return ""

    result = {}
    for sw in rows:
        name = sw["name"]
        yob  = sw.get("year_of_birth", 2014)
        quals = (age_group_quals or {}).get(yob) or default_quals
        if not quals:
            result[name] = {}
            continue
        result[name] = {}
        for (event, course) in quals:
            qd   = quals[(event, course)]
            ev_d = sw["events"].get((event, course))
            pb   = time_to_seconds(ev_d["time_str"]) if ev_d else None

            county_qual = qs(pb, qd.get("county_qt"), qd.get("county_cons"))
            region_qual = qs(pb, qd.get("region_qt"), qd.get("region_cons"))

            result[name][(event, course)] = {
                "county_qual": county_qual,
                "region_qual": region_qual,
            }
    return result


def build_html(rows, run_time, history=None, quals=None, prev_bests=None,
               html_path=None, title=None, peer_histories=None,
               peer_rows_by_yob=None, age_group_quals=None,
               home_url=None, manifest="manifest.json"):
    print("\nBuilding HTML report...")

    # Combine main group with all peer age groups
    all_rows = list(rows)
    for _pr in (peer_rows_by_yob or {}).values():
        all_rows += list(_pr)

    # Primary YOB: the year of birth of the highlighted (is_margot) swimmer
    primary_yob = next(
        (r.get("year_of_birth", 2014) for r in rows if r.get("is_margot")),
        rows[0].get("year_of_birth", 2014) if rows else 2014
    )

    # Unique stroke names in order
    stroke_names = list(dict.fromkeys(name for name, _ in EVENTS))

    def _fam_slug(ev_name): return slug(ev_name.split()[-1])
    def _dist_key(ev_name):
        d = int(ev_name.split()[0])
        return "50m" if d == 50 else ("100m" if d == 100 else "200mp")
    _DIST_LABEL = {"50m": "50m", "100m": "100m", "200mp": "200m+"}

    stroke_groups = []
    i = 0
    while i < len(EVENTS):
        name = EVENTS[i][0]
        span = sum(1 for e in EVENTS[i:] if e[0] == name)
        stroke_groups.append((name, span))
        i += span

    # Columns where no swimmer has a time — used for "hide empty" toggle
    cols_with_times = set()
    for row in all_rows:
        for (ev, course), ev_data in row["events"].items():
            if ev_data and ev_data.get("time_str"):
                cols_with_times.add((slug(ev), slug(course)))
    empty_col_keys = {
        (slug(ev), slug(course))
        for ev, course in EVENTS
        if (slug(ev), slug(course)) not in cols_with_times
    }

    per_swimmer_quals = _compute_swimmer_quals(all_rows, quals, age_group_quals) if (quals or age_group_quals) else {}

    def rank_class(rank):
        return {1: "rank-1", 2: "rank-2", 3: "rank-3"}.get(rank, "")

    # Build stroke header row — add data-stroke for JS filtering
    # ── Season summary panel ──────────────────────────────────────────────────
    margot_row = next((r for r in rows if r["is_margot"]), None)
    if margot_row and quals:
        q_county, c_county, q_region, c_region, first_club = [], [], [], [], []
        # Group by event name — SC+LC qual for same event counts as 1
        _cty_by_ev, _reg_by_ev = {}, {}
        for (ev, course), qdata in quals.items():
            if course == "LC\u2192SC": continue
            for _grade, _store in [
                (qdata.get("county_qual", ""), _cty_by_ev),
                (qdata.get("region_qual",  ""), _reg_by_ev),
            ]:
                if _grade in ("Q", "C"):
                    _store.setdefault(ev, {}).setdefault(_grade, []).append(course)
        def _ev_label(ev, statuses, grade):
            _cs = "/".join(sorted(statuses[grade]))
            return f"{ev} ({_cs})"
        for ev, statuses in _cty_by_ev.items():
            if "Q" in statuses:   q_county.append(_ev_label(ev, statuses, "Q"))
            elif "C" in statuses: c_county.append(_ev_label(ev, statuses, "C"))
        for ev, statuses in _reg_by_ev.items():
            if "Q" in statuses:   q_region.append(_ev_label(ev, statuses, "Q"))
            elif "C" in statuses: c_region.append(_ev_label(ev, statuses, "C"))
        for (ev, course), ev_data in margot_row["events"].items():
            if course != "LC→SC" and ev_data.get("club_rank") == 1:
                first_club.append(f"{ev} {course}")

        def pill(items, bg, label, always=False, pid=""):
            if not items and not always: return ""
            tip = ", ".join(items) if items else "None"
            id_attr = f'id="{pid}" ' if pid else ""
            return (f'<span {id_attr}style="background:{bg};padding:3px 10px;border-radius:12px;'
                    f'font-size:11px;font-weight:600;cursor:default;opacity:{"1" if items else "0.45"}" title="{tip}">'
                    f'{label}: {len(items)}</span>')

        default_first = next((r["name"].split()[0] for r in rows if r.get("is_margot")), "")
        summary_html = (
            '<div style="display:flex;flex-wrap:wrap;gap:8px;margin-bottom:12px;align-items:center">'
            + f'<span id="qual-title" style="font-size:12px;font-weight:700;color:#1a3a5c;margin-right:4px">{default_first} Qualification Status:</span>'
            + pill(first_club, "#d4edda", "🥇 Club 1st")
            + pill(q_county,   "#c8e6ff", "County Q",              pid="qual-cq")
            + pill(c_county,   "#e8f5e9", "County C",              pid="qual-cc")
            + pill(q_region,   "#fce4ec", "Regional Q", always=True, pid="qual-rq")
            + pill(c_region,   "#fff3e0", "Regional C", always=True, pid="qual-rc")
            + '</div>'
        )
    else:
        summary_html = ""

    # ── Progress arrows bar — shown if any event's most recent swim is a PB ────
    def _is_pb_at_latest(ev_name, course_):
        ev_data   = (margot_row or {}).get("events", {}).get((ev_name, course_), {})
        ev_hist   = (history or {}).get((ev_name, course_), [])
        if not ev_data or not ev_hist: return False
        def _d(s):
            try: p=s["date"].split("/"); return (int("20"+p[2]),int(p[1]),int(p[0]))
            except: return (0,0,0)
        most_recent_secs = time_to_seconds(max(ev_hist, key=_d)["time"])
        pb_secs          = time_to_seconds(ev_data.get("time_str",""))
        return bool(most_recent_secs and pb_secs and abs(most_recent_secs - pb_secs) < 0.01)

    has_arrows = bool(margot_row and history and any(
        _is_pb_at_latest(ev, c)
        for (ev, c) in (margot_row["events"] if margot_row else {})
        if c != "LC→SC"
    ))
    progress_bar_html = (
        '<div id="progress-bar" style="margin-bottom:10px;display:flex;align-items:center;gap:12px">'
        '<span style="font-size:11px;color:#27ae60;font-weight:600">▲ PB at most recent meet</span>'
        '<button onclick="toggleArrows(this)" '
        'style="font-size:10px;padding:2px 10px;border:1px solid #ccc;border-radius:3px;'
        'background:#f9f9f9;cursor:pointer">Hide markers</button>'
        '</div>'
    ) if has_arrows else ""

    stroke_hdr = "".join(
        f'<th colspan="{span + 1}" class="stroke-header" data-stroke="{slug(name)}" data-family="{_fam_slug(name)}" data-dist="{_dist_key(name)}">{name}</th>'
        for name, span in stroke_groups
    )

    # JS slug→name map for updateGaps()
    stroke_names_js = "{" + ",".join(
        f'"{slug(n)}":"{n}"'
        for n in dict.fromkeys(n for n, c in EVENTS if c != "LC→SC")
    ) + "}"

    # Build course header row — add data-stroke and data-coltype
    course_hdr_parts = []
    _col_offset = 0  # extra cols added for Best headers
    for j, (name, course) in enumerate(EVENTS):
        empty_attr = ' data-empty="1"' if (slug(name), slug(course)) in empty_col_keys else ''
        course_hdr_parts.append(
            f'<th class="{"conv-header" if course == "LC→SC" else "course-header"}" '
            f'data-col="{3+j+_col_offset}" data-stroke="{slug(name)}" data-family="{_fam_slug(name)}" data-dist="{_dist_key(name)}" data-coltype="{slug(course)}"'
            f'{empty_attr} onclick="sortTable(this)">{course}</th>'
        )
        if course == "LC→SC":
            _col_offset += 1
            course_hdr_parts.append(
                f'<th class="best-header" '
                f'data-col="{3+j+_col_offset}" data-stroke="{slug(name)}" data-family="{_fam_slug(name)}" data-dist="{_dist_key(name)}" data-coltype="best"'
                f' onclick="sortTable(this)">Best</th>'
            )
    course_hdr = "".join(course_hdr_parts)

    # ── Build leader-times lookup: fastest time per (event, course) ──────────
    leader_times = {}  # (event, course) → (time_str, time_secs, swimmer_name)
    for row in rows:
        for (ev, course), ev_data in row["events"].items():
            if course == "LC→SC" or not ev_data.get("time_str"):
                continue
            if ev_data.get("club_rank") == 1:
                leader_times[(ev, course)] = (
                    ev_data["time_str"],
                    time_to_seconds(ev_data["time_str"]),
                    row["name"],
                )

    def gap_str(my_secs, leader_secs):
        """Return '+3.65s off lead' or '' if already leading."""
        if not my_secs or not leader_secs or my_secs <= leader_secs:
            return ""
        diff = round(my_secs - leader_secs, 2)
        if diff >= 60:
            m = int(diff // 60); s = diff - m * 60
            return f"+{m}:{s:04.1f} off lead"
        return f"+{diff:.2f}s off lead"

    # Pre-compute per-swimmer qual event lists for badge tooltips
    _sw_qual_lists = {}
    if per_swimmer_quals:
        for _sw_name, _sw_q in per_swimmer_quals.items():
            _L = {"CQ": [], "CC": [], "RQ": [], "RC": []}
            for (_ev, _c), _qd in _sw_q.items():
                if _c == "LC→SC":
                    continue
                _lab = f"{_ev} {_c}"
                if _qd.get("county_qual") == "Q":   _L["CQ"].append(_lab)
                elif _qd.get("county_qual") == "C": _L["CC"].append(_lab)
                if _qd.get("region_qual") == "Q":   _L["RQ"].append(_lab)
                elif _qd.get("region_qual") == "C": _L["RC"].append(_lab)
            _sw_qual_lists[_sw_name] = _L

    # Build body rows — add data-stroke and data-coltype to every event cell
    body_rows = []
    for swimmer in all_rows:
        is_margot   = swimmer["is_margot"]
        has_history = swimmer.get("has_history", True)
        yob         = swimmer.get("year_of_birth", 2014)
        tr_class    = 'class="margot"' if is_margot else ""
        _safe_name  = swimmer["name"].replace("'", "\\'")
        name_cell   = (
            f'<td class="name" onclick="showPBSummary(\'{_safe_name}\')" '
            f'style="cursor:pointer" title="View {swimmer["name"]}\'s PB summary">'
            f'{swimmer["name"]}</td>'
        )
        posn_cell   = f'<td class="center posn-cell" style="color:#888;font-size:10px">—</td>'
        cnt_cell    = f'<td class="center">{swimmer["event_count"]}</td>'
        yob_cell    = f'<td class="center yob-cell">{yob}</td>'

        event_cells = []
        for short_name, course in EVENTS:
            ev  = swimmer["events"].get((short_name, course))
            _empty_attr = ' data-empty="1"' if (slug(short_name), slug(course)) in empty_col_keys else ''
            da  = f'data-stroke="{slug(short_name)}" data-family="{_fam_slug(short_name)}" data-dist="{_dist_key(short_name)}" data-coltype="{slug(course)}"{_empty_attr}'
            onclick_attr = ""
            progress_arrow = ""
            swimmer_name   = swimmer["name"]
            if course != "LC→SC" and has_history:
                safe_name = swimmer_name.replace("'", "\\'")
                onclick_attr = (
                    f' onclick="showHistory(\'{safe_name}\',\'{short_name}\',\'{course}\')"'
                    f' style="cursor:pointer"'
                    f' title="Click to see {swimmer_name}\'s {short_name} {course} history"'
                )
                # Progress arrow + improvement vs previous PB
                if ev:
                    if is_margot:
                        event_hist = (history or {}).get((short_name, course), [])
                    else:
                        event_hist = (peer_histories or {}).get(swimmer_name, {}).get(f"{short_name}|{course}", [])
                    if event_hist:
                        def _d(s):
                            try:
                                p = s["date"].split("/")
                                return (int("20"+p[2]), int(p[1]), int(p[0]))
                            except: return (0,0,0)
                        swims_sorted = sorted(event_hist, key=_d)
                        most_recent  = swims_sorted[-1]
                        most_recent_secs = time_to_seconds(most_recent["time"])
                        pb_secs = time_to_seconds(ev["time_str"])
                        if most_recent_secs and pb_secs and abs(most_recent_secs - pb_secs) < 0.01:
                            prev_pb = None
                            for sw in swims_sorted[:-1]:
                                s = time_to_seconds(sw["time"])
                                if s and (prev_pb is None or s < prev_pb):
                                    prev_pb = s
                            if prev_pb and prev_pb > pb_secs:
                                imp = prev_pb - pb_secs
                                imp_str = f"-{imp:.2f}s"
                                progress_arrow = (
                                    f'<span class="progress-arrow" title="PB set at most recent meet">'
                                    f' ▲<span style="font-size:9px;font-weight:normal;opacity:0.75"> {imp_str}</span></span>'
                                )
                            else:
                                progress_arrow = '<span class="progress-arrow" title="PB set at most recent meet"> ▲</span>'

            if course == "LC→SC":
                if ev and ev.get("time_str"):
                    rc   = rank_class(ev["club_rank"]) if ev.get("club_rank") else ""
                    secs = time_to_seconds(ev["time_str"]) or 0
                    cell = (
                        f'<td class="converted {rc}" {da} data-secs="{secs:.3f}">'
                        f'<span class="time">{ev["time_str"]}</span>'
                        f'<br><span class="date">{ev["date_str"]}</span></td>'
                    )
                else:
                    cell = f'<td class="converted empty" {da}>—</td>'
            elif ev:
                rc  = rank_class(ev["club_rank"])
                qual_ind = ""
                if quals:
                    qd = per_swimmer_quals.get(swimmer["name"], {}).get((short_name, course), {})
                    cq = qd.get("county_qual", "")
                    rq = qd.get("region_qual", "")
                    parts = []
                    _ql = _sw_qual_lists.get(swimmer["name"], {})
                    if cq == "Q":
                        _tip = "County QT: " + (", ".join(_ql.get("CQ", [])) or "none")
                        parts.append(f'<span class="qi-q" title="{_tip}">CQ</span>')
                    elif cq == "C":
                        _tip = "County Cons: " + (", ".join(_ql.get("CC", [])) or "none")
                        parts.append(f'<span class="qi-c" title="{_tip}">CC</span>')
                    if rq == "Q":
                        _tip = "Regional QT: " + (", ".join(_ql.get("RQ", [])) or "none")
                        parts.append(f'<span class="qi-rq" title="{_tip}">RQ</span>')
                    elif rq == "C":
                        _tip = "Regional Cons: " + (", ".join(_ql.get("RC", [])) or "none")
                        parts.append(f'<span class="qi-rc" title="{_tip}">RC</span>')
                    if parts:
                        qual_ind = (
                            '<span class="qual-ind">'
                            + '<span class="qi-sep">·</span>'.join(parts)
                            + '</span>'
                        )
                secs = time_to_seconds(ev["time_str"]) or 0
                cell = (
                    f'<td class="event {rc}" {da}{onclick_attr} data-secs="{secs:.3f}">'
                    f'<span class="time">{ev["time_str"]}</span>'
                    f'{progress_arrow}'
                    f'<br><span class="date">{ev["date_str"]}</span>'
                    f'{qual_ind}</td>'
                )
            else:
                cell = f'<td class="event empty" {da}{onclick_attr}>—</td>'
            event_cells.append(cell)
            # Append Best Time cell after LC→SC
            if course == "LC→SC":
                _sc_ev   = swimmer["events"].get((short_name, "SC"))
                _conv_ev = swimmer["events"].get((short_name, "LC→SC"))
                _sc_s   = time_to_seconds(_sc_ev["time_str"])   if (_sc_ev   and _sc_ev.get("time_str"))   else None
                _conv_s = time_to_seconds(_conv_ev["time_str"]) if (_conv_ev and _conv_ev.get("time_str")) else None
                _da_b = f'data-stroke="{slug(short_name)}" data-family="{_fam_slug(short_name)}" data-dist="{_dist_key(short_name)}" data-coltype="best"'
                if _sc_s is not None or _conv_s is not None:
                    if _sc_s is not None and (_conv_s is None or _sc_s <= _conv_s):
                        _bt, _bsrc, _bs, _bd = _sc_ev["time_str"], "SC", _sc_s, _sc_ev.get("date_str","")
                    else:
                        _bt, _bsrc, _bs, _bd = _conv_ev["time_str"], "LC→SC", _conv_s, _conv_ev.get("date_str","")
                    _sc = "#27ae60" if _bsrc == "SC" else "#2e75b6"
                    _bcell = (
                        f'<td class="event best-cell" {_da_b} data-secs="{_bs:.3f}">'
                        f'<span class="time">{_bt}</span>'
                        f'<br><span style="font-size:9px;font-weight:700;color:{_sc}">{_bsrc}</span>'
                        f'<br><span class="date">{_bd}</span>'
                        f'</td>'
                    )
                else:
                    _bcell = f'<td class="event empty best-cell" {_da_b}>—</td>'
                event_cells.append(_bcell)

        yob_data      = f' data-yob="{yob}"'
        swimmer_data  = f' data-swimmer="{swimmer["name"].replace(chr(34), "&quot;")}"'
        peer_hidden   = ' style="display:none"' if yob != primary_yob else ""
        body_rows.append(
            f'<tr {tr_class}{yob_data}{swimmer_data}{peer_hidden}>{name_cell}{posn_cell}{yob_cell}{"".join(event_cells)}{cnt_cell}</tr>'
        )

    # Settings panel — stroke checkboxes + column-type checkboxes + age group toggles
    _families = list(dict.fromkeys(_fam_slug(n) for n, _ in EVENTS))
    _family_labels = {_fam_slug(n): n.split()[-1] for n, _ in EVENTS}
    family_checks = "".join(
        f'<label><input type="checkbox" class="family-cb" data-family="{f}" checked> {_family_labels[f]}</label>'
        for f in _families
    )
    dist_checks = "".join(
        f'<label><input type="checkbox" class="dist-cb" data-dist="{d}" checked> {_DIST_LABEL[d]}</label>'
        for d in ["50m", "100m", "200mp"]
        if any(_dist_key(n) == d for n, _ in EVENTS)
    )
    coltype_checks = "".join(
        f'<label><input type="checkbox" class="coltype-cb" data-coltype="{slug(c)}" checked> {c}</label>'
        for c in ["SC", "LC", "LC→SC", "Best"]
    )
    # YOB → human label mapping (extended to cover all possible cohorts)
    _yob_labels = {
        2017: "8-year-olds",  2016: "9-year-olds",  2015: "10-year-olds",
        2014: "11-year-olds", 2013: "12-year-olds",  2012: "13-year-olds",
        # Legacy labels used when Margot was 12, kept for backward compat
        # Overridden per-page by primary_yob logic below
    }
    # Recalibrate labels relative to the page's primary YOB so "primary = 12" always
    # holds true for Margot's page, and "primary = correct age" for other pages.
    # Strategy: compute age from primary_yob and derive sibling cohort labels.
    def _age_label(yob):
        import datetime
        current_year = datetime.date.today().year
        age = current_year - yob  # approximate — good enough for label purposes
        return f"{age}-year-olds"
    _yob_labels = {yob: _age_label(yob) for yob in range(2010, 2020)}

    # Build age group checkboxes — ages 8-9 (YOBs 2017+2018) share one checkbox
    _GROUPED_YOBS = (2017, 2018)
    _row_yobs = sorted(set(r.get("year_of_birth") for r in all_rows if r.get("year_of_birth")))
    _cy = datetime.now().year
    age_group_parts = []
    _handled_young = False
    for _yob in _row_yobs:
        if _yob in _GROUPED_YOBS:
            if not _handled_young:
                _checked = "checked " if primary_yob in _GROUPED_YOBS else ""
                age_group_parts.append(
                    f'<label><input type="checkbox" {_checked}onchange="toggleAgeGroup(\'2017\',this.checked);toggleAgeGroup(\'2018\',this.checked)"> '
                    f'Ages 8–9 (born 2017–2018)</label>'
                )
                _handled_young = True
            continue
        _checked = "checked " if _yob == primary_yob else ""
        _age = _cy - _yob
        age_group_parts.append(
            f'<label><input type="checkbox" {_checked}onchange="toggleAgeGroup(\'{_yob}\',this.checked)"> '
            f'Age {_age} (born {_yob})</label>'
        )
    age_group_section = (
        '<div class="settings-section"><h3>Age groups</h3>'
        f'<div class="checks">{"".join(age_group_parts)}</div></div>'
    )

    # Swimmer dropdown — grouped by age group
    _groups: dict = {}
    for _r in all_rows:
        _groups.setdefault(_r.get("year_of_birth", 2014), []).append(_r["name"])
    _opts = '<option value="">All swimmers</option>'
    for _yob in sorted(_groups.keys()):
        _opts += f'<optgroup label="{_yob_labels.get(_yob, str(_yob))}">'
        for _n in _groups[_yob]:
            _safe = _n.replace('"', '&quot;')
            _opts += f'<option value="{_safe}">{_n}</option>'
        _opts += '</optgroup>'
    swimmer_filter_section = (
        '<div class="settings-section"><h3>Swimmer</h3>'
        f'<select class="swimmer-select" onchange="filterSwimmer(this.value)" '
        f'style="font-size:11px;padding:3px 6px;border:1px solid #ccc;border-radius:4px;'
        f'cursor:pointer;touch-action:manipulation">{_opts}</select></div>'
    )

    # Precompute JS yobLabels string for injection into the template
    yob_labels_js = ",".join(
        f"'{y}':'{_age_label(y)}'" for y in range(2010, 2020)
    )

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{title or "Brompton SC Club Rankings — Female 2014"}</title>
<link rel="manifest" href="{manifest}">
<link rel="apple-touch-icon" href="icon-192.png">
<style>
  *{{box-sizing:border-box}}
  body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;font-size:12px;background:#f5f5f5;margin:0;padding:16px;color:#333}}
  h1{{font-size:18px;font-weight:600;margin:0 0 4px;color:#1a3a5c}}
  .meta{{font-size:11px;color:#888;margin-bottom:12px}}
  /* Settings panel */
  .settings{{background:white;border:1px solid #ddd;border-radius:6px;padding:12px 16px;margin-bottom:14px;box-shadow:0 1px 3px rgba(0,0,0,.08)}}
  .settings-toggle{{font-weight:600;font-size:12px;color:#1a3a5c;cursor:pointer;background:none;border:none;padding:0;display:flex;align-items:center;gap:6px;touch-action:manipulation;-webkit-tap-highlight-color:transparent}}
  .settings-body{{display:flex;gap:32px;margin-top:12px;flex-wrap:wrap}}
  .settings-section h3{{font-size:11px;font-weight:600;color:#555;text-transform:uppercase;letter-spacing:.5px;margin:0 0 8px}}
  .settings-section .checks{{display:flex;flex-wrap:wrap;gap:6px 16px}}
  .settings-section label{{font-size:11px;display:flex;align-items:center;gap:4px;cursor:pointer;white-space:nowrap;touch-action:manipulation}}
  .settings-section label:hover{{color:#1a3a5c}}
  .select-btns{{display:flex;gap:6px;margin-top:6px}}
  .select-btns button{{font-size:10px;padding:2px 8px;border:1px solid #ccc;border-radius:3px;background:#f9f9f9;cursor:pointer;touch-action:manipulation;-webkit-tap-highlight-color:transparent}}
  .select-btns button:hover{{background:#eee}}
  /* Table */
  .wrapper{{overflow-x:auto;background:white}}
  table{{border-collapse:collapse;white-space:nowrap;background:white;box-shadow:0 1px 4px rgba(0,0,0,.1)}}
  th,td{{border:1px solid #ddd;padding:5px 8px;text-align:center;vertical-align:middle}}
  th.stroke-header{{background:#2e75b6;color:white;font-size:11px;font-weight:600;padding:6px 4px;position:sticky;top:0;z-index:4}}
  th.course-header{{background:#1f4e79;color:white;font-size:10px;font-weight:600;cursor:pointer;user-select:none;touch-action:manipulation;-webkit-tap-highlight-color:transparent;position:sticky;z-index:3}}
  th.course-header:hover{{background:#163d6b}}
  th.best-header{{background:#6b4f00;color:white;font-size:10px;font-weight:600;cursor:pointer;user-select:none;touch-action:manipulation;-webkit-tap-highlight-color:transparent;position:sticky;z-index:3}}
  th.best-header:hover{{background:#4a3700}}
  td.best-cell{{background:#fffdf0}}
  th.name-header{{background:#1f4e79;color:white;font-size:11px;font-weight:600;text-align:left;position:sticky;left:0;top:0;z-index:5;min-width:180px;cursor:pointer;user-select:none;touch-action:manipulation;-webkit-tap-highlight-color:transparent}}
  th.name-header:hover{{background:#163d6b}}
  th.cnt-header{{background:#1f4e79;color:white;font-size:11px;font-weight:600;min-width:50px;cursor:pointer;user-select:none;touch-action:manipulation;-webkit-tap-highlight-color:transparent;position:sticky;top:0;z-index:4}}
  th.cnt-header:hover{{background:#163d6b}}
  th.sort-asc::after{{content:" ↑"}}
  th.sort-desc::after{{content:" ↓"}}
  td.name{{text-align:left;font-weight:500;position:sticky;left:0;background:white;z-index:1;touch-action:manipulation;-webkit-tap-highlight-color:transparent}}
  td.event{{font-size:11px;min-width:90px;line-height:1.4;vertical-align:top;touch-action:manipulation;-webkit-tap-highlight-color:transparent}}
  td.event .time{{font-weight:600}}
  td.event .rank{{color:#555;font-size:10px}}
  td.event .date{{color:#888;font-size:10px}}
  td.empty{{color:#ccc}}
  th.conv-header{{background:#34495e;color:#ccc;font-size:9px;font-weight:600;cursor:pointer;user-select:none;position:sticky;z-index:3}}
  th.conv-header:hover{{background:#2c3e50}}
  td.converted{{font-size:11px;min-width:90px;line-height:1.4;vertical-align:top;background:#fafafa;color:#555}}
  td.converted .time{{font-weight:500;color:#444}}
  td.converted .rank{{color:#888;font-size:10px}}
  td.converted .date{{color:#aaa;font-size:10px}}
  td.yob-cell{{font-size:11px;color:#888;font-weight:500}}
  td.rank-1{{background:#d4edda}}
  td.rank-2{{background:#fff3cd}}
  td.rank-3{{background:#fde8d8}}
  tr.margot td.name{{font-weight:700}}
  tr.margot td.name::before{{content:'⭐ '}}
  tr:hover td{{filter:brightness(.96)}}
  .col-hidden{{display:none!important}}
  .progress-arrow{{color:#27ae60;font-size:10px;font-weight:700}}
  .progress-arrow.hidden{{display:none}}
  .gap-to-leader{{color:#999;font-size:9px;display:block;margin-top:1px}}
  .gap-next{{font-size:9px;display:block;margin-top:1px;line-height:1.2;font-weight:600}}
  .gap-next.hidden{{display:none}}
  .qual-ind{{display:block;margin-top:2px;font-size:9px;line-height:1.2}}
  .qual-ind.hidden{{display:none}}
  .qi-q{{color:#27ae60;font-weight:700}}
  .qi-c{{color:#e67e22;font-weight:600}}
  .qi-rq{{color:#2980b9;font-weight:700}}
  .qi-rc{{color:#8e44ad;font-weight:600}}
  .qi-sep{{color:#ccc;margin:0 3px}}
  /* ── PB Summary modal ── */
  #pb-modal{{display:none;position:fixed;inset:0;background:rgba(0,0,0,.45);z-index:10000;
    align-items:center;justify-content:center;padding:16px}}
  #pb-inner{{background:white;border-radius:14px;width:100%;max-width:560px;
    max-height:88vh;overflow-y:auto;box-shadow:0 16px 48px rgba(0,0,0,.3)}}
  .pb-hdr{{display:flex;align-items:flex-start;justify-content:space-between;padding:18px 20px 12px}}
  .pb-title{{font-size:16px;font-weight:700;color:#1a3a5c;margin:0;line-height:1.2}}
  .pb-subtitle{{font-size:11px;color:#888;margin-top:3px}}
  .pb-close{{background:none;border:none;font-size:22px;cursor:pointer;color:#aaa;
    padding:0;line-height:1;flex-shrink:0}}
  .pb-close:hover{{color:#333}}
  .pb-table{{width:100%;border-collapse:collapse;font-size:11px;margin-bottom:16px}}
  .pb-table th{{background:#1f4e79;color:white;font-weight:600;padding:6px 10px;
    text-align:center;font-size:10px;letter-spacing:.3px}}
  .pb-table th.pb-th-event{{text-align:left;min-width:80px}}
  .pb-table th.pb-th-sc{{background:#1565C0}}
  .pb-table th.pb-th-lc{{background:#2e7d32}}
  .pb-stroke-hdr td{{background:#e8edf5;color:#1a3a5c;font-weight:700;font-size:10px;
    text-transform:uppercase;letter-spacing:.6px;padding:5px 10px;border-top:2px solid #c5cfe8}}
  .pb-table td{{padding:5px 10px;border-bottom:1px solid #f0f0f0;vertical-align:middle}}
  .pb-table tr:last-child td{{border-bottom:none}}
  .pb-ev-name{{color:#555;font-weight:500}}
  .pb-t{{font-weight:700;color:#1a3a5c;display:block}}
  .pb-d{{color:#aaa;font-size:10px;display:block;margin-top:1px}}
  .pb-empty{{color:#ddd}}
  .pb-no-times{{font-size:11px;color:#bbb;font-style:italic;padding:8px 20px 16px}}
  /* ── Table sort bar ── */
  /* Small screen tweaks — keep standard table view, just tighten layout */
  @media(max-width:767px){{
    body{{padding:10px}}
    .settings{{padding:10px 12px}}
    .settings-body{{flex-direction:column;gap:16px}}
    h1{{font-size:15px}}
  }}
</style>
<script>
(function(){{
  /* ── Sort ── */
  function cellValue(td, asc){{
    var raw=(td.innerText||td.textContent).trim();
    if(!raw||raw==='—') return asc?Infinity:-Infinity;
    // Extract just the time — take first line, strip ▲ and any trailing non-numeric content
    var token=raw.split('\\n')[0].split('▲')[0].trim();
    var m=token.match(/(\\d+):(\\d+\\.\\d+)/);
    if(m) return parseFloat(m[1])*60+parseFloat(m[2]);
    var n=parseFloat(token);
    if(!isNaN(n)) return n;
    return raw.toLowerCase();
  }}
  window.sortTable=function(th){{
    var table=th.closest('table'),tbody=table.querySelector('tbody');
    var colIdx=parseInt(th.getAttribute('data-col'),10);
    var asc=true;  // always sort fastest first
    table.querySelectorAll('th').forEach(function(h){{h.classList.remove('sort-asc','sort-desc');}});
    th.classList.add('sort-asc');
    var rows=Array.from(tbody.querySelectorAll('tr'));
    rows.sort(function(a,b){{
      var va=cellValue(a.cells[colIdx],asc),vb=cellValue(b.cells[colIdx],asc);
      if(va===vb) return 0;
      if(typeof va==='string'&&typeof vb==='string') return asc?va.localeCompare(vb):vb.localeCompare(va);
      return asc?va-vb:vb-va;
    }});
    rows.forEach(function(r){{tbody.appendChild(r);}});

    updatePosn(tbody,colIdx);
  }};

  function updatePosn(tbody,colIdx){{
    var allRows=Array.from(tbody.querySelectorAll('tr'));
    var visibleRows=allRows.filter(function(r){{return r.style.display!=='none';}});
    var timedRows=visibleRows.filter(function(r){{
      var c=r.cells[colIdx];
      var v=c?(c.innerText||c.textContent).trim():'';
      return v!==''&&v!=='—';
    }});
    allRows.forEach(function(r){{
      var pCell=r.cells[1];
      if(!pCell) return;
      if(r.style.display==='none')return;
      var c=r.cells[colIdx];
      var v=c?(c.innerText||c.textContent).trim():'';
      var hasTime=v!==''&&v!=='—';
      if(hasTime){{
        pCell.textContent=timedRows.indexOf(r)+1;
        pCell.style.cssText='color:#555;font-weight:600;text-align:center;font-size:10px';
      }}else{{
        pCell.textContent='—';
        pCell.style.cssText='color:#ccc;font-weight:normal;text-align:center;font-size:10px';
      }}
    }});
  }}

  /* ── Filter ── */
  var _hideEmpty=false;
  window.toggleEmptyCols=function(btn){{
    _hideEmpty=!_hideEmpty;
    btn.textContent=_hideEmpty?'Show all columns':'Hide empty columns';
    btn.style.background=_hideEmpty?'#2e75b6':'white';
    btn.style.color=_hideEmpty?'white':'#2e75b6';
    applyFilters();
  }};

  function applyFilters(){{
    var table=document.querySelector('table');
    // Collect active families, distances, and coltypes
    var families={{}},dists={{}},coltypes={{}};
    document.querySelectorAll('.family-cb').forEach(function(cb){{families[cb.dataset.family]=cb.checked;}});
    document.querySelectorAll('.dist-cb').forEach(function(cb){{dists[cb.dataset.dist]=cb.checked;}});
    document.querySelectorAll('.coltype-cb').forEach(function(cb){{coltypes[cb.dataset.coltype]=cb.checked;}});

    // Show/hide course headers (row 2) and body cells
    table.querySelectorAll('th[data-family][data-dist][data-coltype]').forEach(function(th){{
      var f=th.dataset.family,d=th.dataset.dist,c=th.dataset.coltype;
      var hidden=!families[f]||!dists[d]||!coltypes[c]||(_hideEmpty&&th.dataset.empty==='1');
      th.classList.toggle('col-hidden',hidden);
    }});
    table.querySelectorAll('td[data-family][data-dist]').forEach(function(td){{
      var f=td.dataset.family,d=td.dataset.dist,c=td.dataset.coltype;
      var hidden=!families[f]||!dists[d]||!coltypes[c]||(_hideEmpty&&td.dataset.empty==='1');
      td.classList.toggle('col-hidden',hidden);
    }});

    // Show/hide stroke group headers (row 1) — hide when family or dist unchecked or all children hidden
    table.querySelectorAll('th.stroke-header').forEach(function(th){{
      var f=th.dataset.family,d=th.dataset.dist;
      if(!families[f]||!dists[d]){{ th.classList.add('col-hidden'); return; }}
      var children=Array.from(table.querySelectorAll('th[data-family="'+f+'"][data-dist="'+d+'"][data-coltype]'));
      var visible=children.filter(function(c){{return !c.classList.contains('col-hidden');}});
      th.classList.toggle('col-hidden',visible.length===0);
      if(visible.length>0) th.setAttribute('colspan',visible.length);
    }});
    if(typeof updateGaps==='function') updateGaps();
  }}

  /* ── Gap to next standard ── */
  var _STROKE_NAMES={stroke_names_js};
  function _nextQual(secs,qd){{
    var cc=qd.county_cons,cq=qd.county_qt,rc=qd.region_cons,rq=qd.region_qt;
    if(rq&&secs<=rq) return null;
    if(rc&&secs<=rc) return rq?{{label:'RQ',gap:secs-rq,color:'#8e44ad'}}:null;
    if(cq&&secs<=cq) return rc?{{label:'RC',gap:secs-rc,color:'#16a085'}}:null;
    if(cc&&secs<=cc) return cq?{{label:'CQ',gap:secs-cq,color:'#e67e22'}}:null;
    return cc?{{label:'CC',gap:secs-cc,color:'#e67e22'}}:null;
  }}
  window.updateGaps=function(){{
    document.querySelectorAll('.gap-next').forEach(function(el){{el.remove();}});
    var gapsHidden=document.querySelector('#progress-bar button')&&document.querySelector('#progress-bar button').textContent==='Show markers';
    document.querySelectorAll('tbody tr').forEach(function(row){{
      row.querySelectorAll('td[data-stroke][data-coltype]').forEach(function(td){{
        var ct=td.getAttribute('data-coltype');if(ct==='lc-sc'||ct==='best') return;
        var secs=parseFloat(td.getAttribute('data-secs'));if(!secs||isNaN(secs)) return;
        var stroke=_STROKE_NAMES[td.getAttribute('data-stroke')];if(!stroke) return;
        var course=ct==='sc'?'SC':'LC';
        var qd=(typeof QT_DATA!=='undefined'&&QT_DATA[stroke+'|'+course])||{{}};
        var info=_nextQual(secs,qd);if(!info) return;
        var sp=document.createElement('span');
        sp.className='gap-next'+(gapsHidden?' hidden':'');
        sp.style.color=info.color;
        sp.textContent='-'+Math.abs(info.gap).toFixed(2)+'s → '+info.label;
        td.appendChild(sp);
      }});
    }});
  }};

  window.toggleArrows=function(btn){{
    var hidden=btn.textContent==='Hide markers';
    document.querySelectorAll('.progress-arrow,.qual-ind,.gap-next').forEach(function(el){{el.classList.toggle('hidden',hidden);}});
    btn.textContent=hidden?'Show markers':'Hide markers';
  }};

  function rerank(){{
    var table=document.querySelector('table');
    if(!table) return;
    var tbody=table.querySelector('tbody');
    var allRows=Array.from(tbody.querySelectorAll('tr'));
    var visibleRows=allRows.filter(function(r){{return r.style.display!=='none';}});
    // Group ALL rows by YOB — rankings use the full age group, not just visible subset
    var byYob={{}};
    allRows.forEach(function(r){{
      var y=r.dataset.yob||'{primary_yob}';
      if(!byYob[y]) byYob[y]=[];
      byYob[y].push(r);
    }});
    table.querySelectorAll('th[data-stroke][data-coltype]').forEach(function(th){{
      var colIdx=parseInt(th.getAttribute('data-col'),10);
      // Clear rank classes from visible rows only
      visibleRows.forEach(function(r){{
        var td=r.cells[colIdx];
        if(td) td.classList.remove('rank-1','rank-2','rank-3');
      }});
      // Rank within each age group; apply class only if the row is currently visible
      Object.keys(byYob).forEach(function(yob){{
        var cells=byYob[yob].map(function(r){{return r.cells[colIdx];}})
          .filter(function(td){{return td&&td.dataset.secs&&parseFloat(td.dataset.secs)>0;}});
        cells.slice().sort(function(a,b){{return parseFloat(a.dataset.secs)-parseFloat(b.dataset.secs);}})
          .slice(0,3).forEach(function(td,i){{
            if(td.closest('tr').style.display!=='none')
              td.classList.add(['rank-1','rank-2','rank-3'][i]);
          }});
      }});
    }});
    var sortedTh=table.querySelector('th.sort-asc');
    if(sortedTh) updatePosn(tbody,parseInt(sortedTh.getAttribute('data-col'),10));
  }}

  function ageGroupChecked(yob){{
    var result=true;
    document.querySelectorAll('input[onchange*="toggleAgeGroup"]').forEach(function(cb){{
      var ms=cb.getAttribute('onchange').match(/'(\d+)'/g);
      if(ms&&ms.some(function(m){{return m.replace(/'/g,'')===String(yob);}})) result=cb.checked;
    }});
    return result;
  }}

  function updateSwimmerDropdown(){{
    var sel=document.querySelector('.swimmer-select');
    if(!sel) return;
    var cur=sel.value;
    var yobLabels={{{yob_labels_js}}};
    Array.from(sel.querySelectorAll('optgroup')).forEach(function(og){{
      var label=og.getAttribute('label');
      var visible=Object.keys(yobLabels).some(function(y){{return yobLabels[y]===label&&ageGroupChecked(y);}});
      og.style.display=visible?'':'none';
      if(!visible&&cur){{
        Array.from(og.querySelectorAll('option')).forEach(function(opt){{
          if(opt.value===cur){{sel.value='';filterSwimmer('');}}
        }});
      }}
    }});
  }}

  var _qualStrokes={{'50-back':'50 Back','100-back':'100 Back','200-back':'200 Back',
    '50-fly':'50 Fly','100-fly':'100 Fly','200-fly':'200 Fly',
    '50-free':'50 Free','100-free':'100 Free','200-free':'200 Free',
    '400-free':'400 Free','800-free':'800 Free',
    '50-breast':'50 Breast','100-breast':'100 Breast','200-breast':'200 Breast',
    '100-im':'100 IM','200-im':'200 IM','400-im':'400 IM'}};
  function _qualInfo(row,cls){{
    var spans=row.querySelectorAll('.'+cls),seen={{}},uniqueEvents=[];
    spans.forEach(function(s){{
      var td=s.closest('td');if(!td)return;
      var ct=td.getAttribute('data-coltype'),nm=_qualStrokes[td.getAttribute('data-stroke')]||'';
      var course=ct==='sc'?'SC':ct==='lc'?'LC':'';
      if(nm&&course){{if(!seen[nm])seen[nm]=[];seen[nm].push(course);}}
    }});
    Object.keys(seen).forEach(function(nm){{
      uniqueEvents.push(nm+' ('+seen[nm].sort().join('/')+')');
    }});
    return{{count:uniqueEvents.length,title:uniqueEvents.join(', ')||'None'}};
  }}
  function _setBadge(id,info){{
    var el=document.getElementById(id);if(!el)return;
    el.textContent=el.textContent.replace(/\d+$/,'')+info.count;
    el.title=info.title;
    el.style.opacity=info.count>0?'1':'0.45';
  }}
  function _updateHeadingAndQuals(name,row){{
    var sn=document.getElementById('page-swimmer-name');
    if(sn) sn.textContent=name?' — '+name:'';
    var qt=document.getElementById('qual-title');
    if(qt) qt.textContent=(name?name.split(' ')[0]:'')+(name?' Qualification Status:':' Qualification Status:');
    if(row){{
      _setBadge('qual-cc',_qualInfo(row,'qi-c'));
      _setBadge('qual-cq',_qualInfo(row,'qi-q'));
      _setBadge('qual-rc',_qualInfo(row,'qi-rc'));
      _setBadge('qual-rq',_qualInfo(row,'qi-rq'));
    }}
  }}
  window.filterSwimmer=function(val){{
    var tbody=document.querySelector('table tbody');
    Array.from(tbody.querySelectorAll('tr')).forEach(function(r){{
      if(!val){{
        r.style.display=ageGroupChecked(r.dataset.yob)?'':'none';
      }}else{{
        r.style.display=(r.dataset.swimmer===val)?'':'none';
      }}
    }});
    rerank();
    if(typeof updateGaps==='function') updateGaps();
    var defName=window._pageDefaultSwimmer||'';
    var row=val?document.querySelector('tbody tr[data-swimmer="'+val+'"]'):
                (defName?document.querySelector('tbody tr[data-swimmer="'+defName+'"]'):null);
    var name=val||(defName||'');
    _updateHeadingAndQuals(name,row);
  }};

  window.toggleAgeGroup=function(yob,checked){{
    document.querySelectorAll('tr[data-yob="'+yob+'"]').forEach(function(r){{
      r.style.display=checked?'':'none';
    }});
    updateSwimmerDropdown();
    rerank();
    if(typeof updateGaps==='function') updateGaps();
  }};

  var _STORAGE_KEY='swim_filter_defaults_{primary_yob}';

  window.saveFilterDefaults=function(btn){{
    var state={{}};
    document.querySelectorAll('.family-cb').forEach(function(cb){{state['f_'+cb.dataset.family]=cb.checked;}});
    document.querySelectorAll('.dist-cb').forEach(function(cb){{state['d_'+cb.dataset.dist]=cb.checked;}});
    document.querySelectorAll('.coltype-cb').forEach(function(cb){{state['c_'+cb.dataset.coltype]=cb.checked;}});
    document.querySelectorAll('input[onchange*="toggleAgeGroup"]').forEach(function(cb){{
      var ms=cb.getAttribute('onchange').match(/'(\d+)'/g);
      if(ms) ms.forEach(function(m){{state['a_'+m.replace(/'/g,'')]=cb.checked;}});
    }});
    var sw=document.querySelector('.swimmer-select');
    if(sw) state['swimmer']=sw.value;
    try{{localStorage.setItem(_STORAGE_KEY,JSON.stringify(state));}}catch(e){{}}
    var orig=btn.textContent;
    btn.textContent='Saved ✓';
    btn.style.background='#2e75b6';btn.style.color='white';
    setTimeout(function(){{btn.textContent=orig;btn.style.background='white';btn.style.color='#2e75b6';}},1500);
  }};

  function loadFilterDefaults(){{
    var raw;
    try{{raw=localStorage.getItem(_STORAGE_KEY);}}catch(e){{return;}}
    if(!raw) return;
    var state=JSON.parse(raw);
    document.querySelectorAll('.family-cb').forEach(function(cb){{
      if('f_'+cb.dataset.family in state) cb.checked=state['f_'+cb.dataset.family];
    }});
    document.querySelectorAll('.dist-cb').forEach(function(cb){{
      if('d_'+cb.dataset.dist in state) cb.checked=state['d_'+cb.dataset.dist];
    }});
    document.querySelectorAll('.coltype-cb').forEach(function(cb){{
      if('c_'+cb.dataset.coltype in state) cb.checked=state['c_'+cb.dataset.coltype];
    }});
    document.querySelectorAll('input[onchange*="toggleAgeGroup"]').forEach(function(cb){{
      var m=cb.getAttribute('onchange').match(/'(\d+)'/);
      if(m&&'a_'+m[1] in state){{
        cb.checked=state['a_'+m[1]];
        var ms=cb.getAttribute('onchange').match(/'(\d+)'/g);
        if(ms) ms.forEach(function(yobStr){{toggleAgeGroup(yobStr.replace(/'/g,''),cb.checked);}});
      }}
    }});
    var sw=document.querySelector('.swimmer-select');
    if(sw&&'swimmer' in state){{sw.value=state['swimmer'];filterSwimmer(state['swimmer']);}}
    applyFilters();
  }};

  window.toggleSettings=function(btn){{
    var body=document.getElementById('settings-body');
    var open=body.style.display!=='none';
    body.style.display=open?'none':'flex';
    btn.textContent=open?'⚙️ Filter columns ▶':'⚙️ Filter columns ▼';
    setTimeout(function(){{if(typeof _initSticky==='function')_initSticky();}},0);
  }};

  window.addEventListener('DOMContentLoaded',function(){{
    loadFilterDefaults();
    // Hide rows for any age group whose checkbox starts unchecked
    document.querySelectorAll('input[onchange*="toggleAgeGroup"]').forEach(function(cb){{
      if(!cb.checked){{
        var ms=cb.getAttribute('onchange').match(/'(\d+)'/g);
        if(ms) ms.forEach(function(m){{toggleAgeGroup(m.replace(/'/g,''),false);}});
      }}
    }});
    rerank();
    if(typeof updateGaps==='function') updateGaps();
    // Collapse filter panel by default on mobile
    if(window.innerWidth<=767){{
      var sb=document.getElementById('settings-body');
      var st=document.querySelector('.settings-toggle');
      if(sb){{sb.style.display='none';}}
      if(st){{st.textContent='⚙️ Filter columns ▶';}}
    }}
    // Handle ?swimmer= URL param when returning from index.html
    var _urlSwimmer=new URLSearchParams(window.location.search).get('swimmer');
    if(_urlSwimmer){{
      var _sw=document.querySelector('.swimmer-select');
      if(_sw){{_sw.value=_urlSwimmer;filterSwimmer(_urlSwimmer);}}
    }}
    document.querySelectorAll('.family-cb,.dist-cb,.coltype-cb').forEach(function(cb){{
      cb.addEventListener('change',applyFilters);
    }});
    // Select all / none helpers
    document.querySelectorAll('.select-btns button').forEach(function(btn){{
      btn.addEventListener('click',function(){{
        var group=btn.closest('.settings-section').querySelectorAll('input[type=checkbox]');
        var val=btn.dataset.val==='all';
        group.forEach(function(cb){{cb.checked=val;}});
        applyFilters();
      }});
    // Meet-name tooltip on PB date spans (looks up HISTORY)
    document.querySelectorAll('td.event[data-secs]').forEach(function(td){{
      var ct=td.getAttribute('data-coltype');
      if(ct==='best'||ct==='lc-sc') return;
      var course=ct==='sc'?'SC':ct==='lc'?'LC':null;
      if(!course||typeof HISTORY==='undefined') return;
      var row=td.closest('tr');if(!row) return;
      var swimmer=row.getAttribute('data-swimmer');if(!swimmer) return;
      var eventName=_qualStrokes[td.getAttribute('data-stroke')];if(!eventName) return;
      var hist=(HISTORY[swimmer]||{{}})[eventName+'|'+course];
      if(!hist||!hist.length) return;
      var pbSecs=parseFloat(td.getAttribute('data-secs'));
      var match=null;
      hist.forEach(function(h){{
        var hs=parseSeconds(h.time);
        if(Math.abs(hs-pbSecs)<0.01) match=h;
      }});
      if(!match||!match.meet) return;
      var dateSp=td.querySelector('.date');if(!dateSp) return;
      dateSp.title=match.meet;
      dateSp.style.textDecorationLine='underline';
      dateSp.style.textDecorationStyle='dotted';
      dateSp.style.textDecorationColor='#aaa';
      dateSp.style.cursor='help';
    }});
    }});
    // Sticky headers: wrapper must scroll vertically so position:sticky works within it
    var _w=document.querySelector('.wrapper');
    var _tr1=document.querySelector('thead tr:first-child');
    var _tr2=document.querySelector('thead tr:nth-child(2)');
    function _initSticky(){{
      if(_w){{
        _w.style.overflowY='auto';
        var top=_w.getBoundingClientRect().top;
        var avail=(window.innerHeight-top-8)+'px';
        _w.style.maxHeight=avail;
        _w.style.minHeight=avail;
      }}
      if(_tr1&&_tr2){{
        var h=_tr1.getBoundingClientRect().height||_tr1.offsetHeight;
        Array.from(_tr2.querySelectorAll('th')).forEach(function(th){{th.style.top=h+'px';}});
      }}
    }}
    _initSticky();
    window.addEventListener('resize',_initSticky);
  }});
}})();
</script>
</head>
<body>
<h1 id="page-h1">{title or "Brompton SC — Female 2013, 14 &amp; 15 — Personal Bests"}<span id="page-swimmer-name"></span></h1>
<p class="meta">Last updated: {run_time} &nbsp;|&nbsp;
  <span style="background:#d4edda;padding:1px 6px;border-radius:3px">1st</span>
  <span style="background:#fff3cd;padding:1px 6px;border-radius:3px">2nd</span>
  <span style="background:#fde8d8;padding:1px 6px;border-radius:3px">3rd</span>
  &nbsp;|&nbsp; Times from <a href="https://www.swimmingresults.org/12months/" target="_blank" style="color:#888;font-size:11px">Swim England</a>
</p>
{summary_html}
{progress_bar_html}

<div class="settings">
  <button class="settings-toggle" onclick="toggleSettings(this)">⚙️ Filter columns ▼</button>
  <div class="settings-body" id="settings-body">
    <div class="settings-section">
      <h3>Strokes</h3>
      <div class="checks">{family_checks}</div>
      <div class="select-btns">
        <button data-val="all">All</button>
        <button data-val="none">None</button>
      </div>
    </div>
    <div class="settings-section">
      <h3>Distances</h3>
      <div class="checks">{dist_checks}</div>
    </div>
    <div class="settings-section">
      <h3>Column type</h3>
      <div class="checks">{coltype_checks}</div>
    </div>
    {age_group_section}
    {swimmer_filter_section}
    <div class="settings-section" style="justify-content:flex-end;align-self:flex-end;margin-left:auto">
      <button onclick="saveFilterDefaults(this)"
        style="font-size:11px;padding:4px 14px;border:1px solid #2e75b6;border-radius:12px;
               background:white;color:#2e75b6;cursor:pointer;font-weight:600;touch-action:manipulation;
               -webkit-tap-highlight-color:transparent">
        Save as default
      </button>
    </div>
  </div>
</div>



<div class="wrapper">
<table>
<thead>
  <tr>
    <th class="name-header" rowspan="2" data-col="0" onclick="sortTable(this)">Swimmer</th>
    <th class="cnt-header" rowspan="2" id="posn-header" style="min-width:36px">Posn</th>
    <th class="cnt-header" rowspan="2" data-col="2" onclick="sortTable(this)" style="min-width:44px">YOB</th>
    {stroke_hdr}
    <th class="cnt-header" rowspan="2" data-col="{3+len(EVENTS)}" onclick="sortTable(this)">Events</th>
  </tr>
  <tr>{course_hdr}</tr>
</thead>
<tbody>{"".join(body_rows)}</tbody>
</table>
</div>
</body>
</html>"""

    # ── Embed history + modal via shared script ─────────────────────────────
    if history:
        import json as _json
        # Highlighted swimmer's history keyed by event|course
        hi_hist = {f"{ev}|{c}": v for (ev, c), v in history.items()}
        # All swimmers: build nested dict {swimmerName: {event|course: swims}}
        all_hist = {}
        hi_name = next((r["name"] for r in rows if r["is_margot"]), "")
        # Merge locally-stored manual swims (times not yet on Swim England)
        _manual_files = {
            "Margot Mandikos": MARGOT_DIR / "margot_manual_swims.json",
            "Ava Mandikos":    AVA_DIR    / "ava_manual_swims.json",
        }
        _manual_path = _manual_files.get(hi_name)
        if _manual_path and _manual_path.exists():
            _manual = _json.loads(_manual_path.read_text())
            for _key, _swims in _manual.items():
                _seen = {(s['date'], s['time']) for s in hi_hist.get(_key, [])}
                for _s in _swims:
                    if (_s['date'], _s['time']) not in _seen:
                        hi_hist.setdefault(_key, []).append(_s)
        # Seed empty history slots with current PB from table data so the PB
        # popup can show events (e.g. 100 IM) that swimmingresults.org has no
        # full-history listing for.
        _hi_row = next((r for r in rows if r.get("is_margot")), None)
        if _hi_row:
            for (ev, course), ev_data in _hi_row.get("events", {}).items():
                if course == "LC→SC":
                    continue
                _key = f"{ev}|{course}"
                if not hi_hist.get(_key) and ev_data and ev_data.get("time_str") and ev_data.get("date_str"):
                    hi_hist[_key] = [{"time": ev_data["time_str"], "date": ev_data["date_str"], "pts": "", "round": "", "meet": ""}]
        all_hist[hi_name] = hi_hist  # highlighted swimmer always included
        if peer_histories:
            for name, phist in peer_histories.items():
                if name != hi_name:
                    all_hist[name] = phist  # already keyed by "event|course"
        hist_js = all_hist
        qt_js = {}
        for (ev, c), d in (quals or {}).items():
            entry = {k: v for k, v in d.items() if k in ("county_qt","county_cons","region_qt","region_cons")}
            lt = leader_times.get((ev, c))
            if lt:
                entry["leader_secs"] = lt[1]
                entry["leader_name"] = lt[2]
            qt_js[f"{ev}|{c}"] = entry
        for (ev, c), (lt_str, lt_secs, lt_name) in leader_times.items():
            key = f"{ev}|{c}"
            if key not in qt_js:
                qt_js[key] = {}
            qt_js[key].setdefault("leader_secs", lt_secs)
            qt_js[key].setdefault("leader_name", lt_name)
        modal_div = """
<style>
@media (orientation:landscape) and (max-height:520px){
  #modal-inner{height:96vh!important;padding:8px 14px!important;min-width:0!important;min-height:0!important}
  #modal-body{flex-direction:row!important;gap:8px!important}
  #modal-chart-wrap{min-height:0!important}
  #modal-chart-canvas-wrap{min-height:140px!important}
  #modal-table-wrap{min-height:0!important;flex:1.2!important}
}
</style>
<div id="history-modal" style="display:none;position:fixed;top:0;left:0;width:100%;height:100%;
  background:rgba(0,0,0,.5);z-index:1000;align-items:center;justify-content:center;">
  <div id="modal-inner" style="background:white;border-radius:8px;padding:20px 24px;
    width:92vw;height:88vh;display:flex;flex-direction:column;
    box-shadow:0 8px 32px rgba(0,0,0,.3);resize:both;overflow:auto;min-width:500px;min-height:400px">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px;flex-shrink:0">
      <h2 id="modal-title" style="margin:0;font-size:15px;color:#1a3a5c"></h2>
      <div style="display:flex;gap:8px;align-items:center">
        <button onclick="toggleFullscreen()" title="Toggle fullscreen"
          style="border:1px solid #ddd;background:#f9f9f9;font-size:13px;cursor:pointer;
            padding:2px 8px;border-radius:3px;color:#555">⤢</button>
        <button onclick="closeModal()"
          style="border:none;background:none;font-size:22px;cursor:pointer;color:#666;line-height:1">&times;</button>
      </div>
    </div>
    <div id="modal-qual-summary" style="margin-bottom:10px;flex-shrink:0"></div>
    <div id="modal-body" style="flex:1;min-height:0;display:flex;flex-direction:column;gap:12px;overflow:hidden">
      <div id="modal-chart-wrap" style="flex:2;min-height:0;display:flex;flex-direction:column">
        <div id="modal-chart-canvas-wrap" style="position:relative;flex:1;min-height:320px;touch-action:none">
          <canvas id="history-chart" style="touch-action:none;user-select:none"></canvas>
        </div>
        <div id="chart-legend" style="display:flex;flex-wrap:wrap;gap:8px 16px;margin-top:8px;font-size:11px;flex-shrink:0"></div>
        <div id="chart-period-btns" style="display:flex;gap:6px;margin-top:8px;flex-shrink:0">
          <button onclick="filterChartPeriod(3,this)"  style="border:1px solid #ccc;border-radius:12px;font-size:11px;padding:3px 10px;cursor:pointer;background:#f0f4f8;color:#444;touch-action:manipulation">3M</button>
          <button onclick="filterChartPeriod(6,this)"  style="border:1px solid #ccc;border-radius:12px;font-size:11px;padding:3px 10px;cursor:pointer;background:#f0f4f8;color:#444;touch-action:manipulation">6M</button>
          <button onclick="filterChartPeriod(12,this)" style="border:1px solid #ccc;border-radius:12px;font-size:11px;padding:3px 10px;cursor:pointer;background:#f0f4f8;color:#444;touch-action:manipulation">1Y</button>
          <button onclick="filterChartPeriod(24,this)" style="border:1px solid #ccc;border-radius:12px;font-size:11px;padding:3px 10px;cursor:pointer;background:#f0f4f8;color:#444;touch-action:manipulation">2Y</button>
          <button onclick="filterChartPeriod(null,this)" style="border:1px solid #ccc;border-radius:12px;font-size:11px;padding:3px 10px;cursor:pointer;background:#f0f4f8;color:#444;touch-action:manipulation;background:#1f4e79;color:white;border-color:#1f4e79">All</button>
        </div>
      </div>
      <div id="modal-table-wrap" style="flex:1;overflow-y:auto;min-height:80px">
        <table id="modal-table" style="width:100%;border-collapse:collapse;font-size:12px"></table>
      </div>
    </div>
  </div>
</div>"""
        cdn = (
            '<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>\n'
            '<script src="https://cdn.jsdelivr.net/npm/chartjs-adapter-date-fns@3.0.0/dist/chartjs-adapter-date-fns.bundle.min.js"></script>'
        )
        pb_modal_html = """
<div id="pb-modal" onclick="if(event.target===this)closePBModal()">
  <div id="pb-inner">
    <div class="pb-hdr">
      <div>
        <div class="pb-title" id="pb-title"></div>
        <div class="pb-subtitle">Personal Bests</div>
      </div>
      <button class="pb-close" onclick="closePBModal()">&times;</button>
    </div>
    <div id="pb-body"></div>
  </div>
</div>
<script>
(function(){
var PB_STROKES=[
  {label:'Backstroke',   events:['50 Back','100 Back','200 Back']},
  {label:'Butterfly',    events:['50 Fly','100 Fly','200 Fly']},
  {label:'Freestyle',    events:['50 Free','100 Free','200 Free','400 Free','800 Free']},
  {label:'Breaststroke', events:['50 Breast','100 Breast','200 Breast']},
  {label:'Ind. Medley',  events:['100 IM','200 IM','400 IM']},
];
function pbSecs(t){
  if(!t) return null;
  var m=t.match(/^(\\d+):(\\d+\\.?\\d*)$/);
  if(m) return parseInt(m[1])*60+parseFloat(m[2]);
  var f=parseFloat(t);
  return isNaN(f)?null:f;
}
function getBest(hist,key){
  var swims=(hist&&hist[key])||[];
  var best=null;
  swims.forEach(function(s){
    var secs=pbSecs(s.time);
    if(secs!==null&&(best===null||secs<best.secs))
      best={time:s.time,date:s.date,secs:secs};
  });
  return best;
}
function cell(b){
  if(!b) return '<td class="pb-empty" style="text-align:center">—</td><td class="pb-empty" style="text-align:center;color:#eee">—</td>';
  return '<td style="text-align:center"><span class="pb-t">'+b.time+'</span></td>'
    +'<td style="text-align:center"><span class="pb-d" style="display:block">'+b.date+'</span></td>';
}
window.showPBSummary=function(name){
  var hist=(typeof HISTORY!=='undefined'&&HISTORY[name])||{};
  document.getElementById('pb-title').textContent=name;
  var hasAny=false;
  var rows='';
  PB_STROKES.forEach(function(stroke){
    var strokeRows='';
    stroke.events.forEach(function(ev){
      var sc=getBest(hist,ev+'|SC');
      var lc=getBest(hist,ev+'|LC');
      if(!sc&&!lc) return;
      hasAny=true;
      var dist=ev.split(' ')[0]+'m';
      strokeRows+='<tr><td class="pb-ev-name">'+dist+'</td>'+cell(sc)+cell(lc)+'</tr>';
    });
    if(strokeRows){
      rows+='<tr class="pb-stroke-hdr"><td colspan="5">'+stroke.label+'</td></tr>'+strokeRows;
    }
  });
  var body='';
  if(hasAny){
    body='<div style="padding:0 16px 16px">'
      +'<table class="pb-table">'
      +'<thead><tr>'
      +'<th class="pb-th-event">Event</th>'
      +'<th class="pb-th-sc" colspan="2">Short Course</th>'
      +'<th class="pb-th-lc" colspan="2">Long Course</th>'
      +'</tr>'
      +'<tr>'
      +'<th class="pb-th-event"></th>'
      +'<th class="pb-th-sc">Time</th><th class="pb-th-sc">Date</th>'
      +'<th class="pb-th-lc">Time</th><th class="pb-th-lc">Date</th>'
      +'</tr></thead>'
      +'<tbody>'+rows+'</tbody>'
      +'</table></div>';
  } else {
    body='<div class="pb-no-times">No times recorded yet.</div>';
  }
  document.getElementById('pb-body').innerHTML=body;
  document.getElementById('pb-modal').style.display='flex';
};
window.closePBModal=function(){
  document.getElementById('pb-modal').style.display='none';
};
document.addEventListener('keydown',function(e){if(e.key==='Escape')closePBModal();});
})();
</script>"""
        html = html.replace(
            "</body>",
            cdn + "\n" + modal_div + "\n"
            + _modal_script(_json.dumps(hist_js), _json.dumps(qt_js))
            + pb_modal_html
            + "\n</body>"
        )


    # Personalisation: read bsc_swimmer from localStorage and move ⭐ row + update qual title
    if html_path:
        import re as _re
        this_filename = html_path.name
        default_name  = next((r["name"] for r in rows if r.get("is_margot")), "")
        default_first = default_name.split()[0] if default_name else ""
        year_m        = _re.search(r'\d{2}/\d{2}/(\d{4})', run_time)
        qual_year     = str(int(year_m.group(1)) + 1) if year_m else "2027"
        personalise_js = (
            "<script>\n(function(){\n"
            f"  window._pageDefaultSwimmer = '{default_name}';\n"
            f"  var fullName, firstName = '{default_first}';\n"
            "  try { fullName = localStorage.getItem('bsc_swimmer'); } catch(e){}\n"
            "  if(fullName) firstName = fullName.split(' ')[0];\n"
            "  var savedUrl = '';\n"
            "  try { savedUrl = localStorage.getItem('bsc_swimmer_url') || ''; } catch(e){}\n"
            f"  if(savedUrl && savedUrl !== '{this_filename}') {{ fullName = null; firstName = '{default_first}'; }}\n"
            f"  try {{ localStorage.setItem('bsc_swimmer', '{default_name}'); localStorage.setItem('bsc_swimmer_url', '{this_filename}'); }} catch(e){{}}\n"
            "  try {\n"
            "    var highlighted = document.querySelector('tr.margot');\n"
            "    if(highlighted) highlighted.classList.remove('margot');\n"
            "    var row = fullName ? document.querySelector('tr[data-swimmer=\"'+fullName+'\"]') : null;\n"
            "    if(row) row.classList.add('margot');\n"
            "    var qt = document.getElementById('qual-title');\n"
            "    if(qt) qt.textContent = firstName + ' Qualification Status:';\n"
            "    var sn = document.getElementById('page-swimmer-name');\n"
            "    if(sn) sn.textContent = fullName ? ' — ' + fullName : '';\n"
            "    if(row){\n"
            "      var strokes={'50-back':'50 Back','100-back':'100 Back','200-back':'200 Back',"
            "'50-fly':'50 Fly','100-fly':'100 Fly','200-fly':'200 Fly',"
            "'50-free':'50 Free','100-free':'100 Free','200-free':'200 Free',"
            "'400-free':'400 Free','800-free':'800 Free',"
            "'50-breast':'50 Breast','100-breast':'100 Breast','200-breast':'200 Breast',"
            "'100-im':'100 IM','200-im':'200 IM','400-im':'400 IM'};\n"
            "      function qualInfo(cls){\n"
            "        var spans=row.querySelectorAll('.'+cls),seen={},uniqueEvents=[];\n"
            "        spans.forEach(function(s){\n"
            "          var td=s.closest('td');if(!td)return;\n"
            "          var ct=td.getAttribute('data-coltype'),nm=strokes[td.getAttribute('data-stroke')]||'';\n"
            "          var course=ct==='sc'?'SC':ct==='lc'?'LC':'';\n"
            "          if(nm&&course){if(!seen[nm])seen[nm]=[];seen[nm].push(course);}\n"
            "        });\n"
            "        Object.keys(seen).forEach(function(nm){uniqueEvents.push(nm+' ('+seen[nm].sort().join('/')+')');});\n"
            "        return{count:uniqueEvents.length,title:uniqueEvents.join(', ')||'None'};\n"
            "      }\n"
            "      function setBadge(id,info){\n"
            "        var el=document.getElementById(id);if(!el)return;\n"
            r"        el.textContent=el.textContent.replace(/\d+$/,'')+info.count;" + "\n"
            "        el.title=info.title;\n"
            "        el.style.opacity=info.count>0?'1':'0.45';\n"
            "      }\n"
            "      setBadge('qual-cc',qualInfo('qi-c'));\n"
            "      setBadge('qual-cq',qualInfo('qi-q'));\n"
            "      setBadge('qual-rc',qualInfo('qi-rc'));\n"
            "      setBadge('qual-rq',qualInfo('qi-rq'));\n"
            "    }\n"
            "    var sw=document.querySelector('.swimmer-select');\n"
            "    if(sw&&fullName){sw.value=fullName;}\n"
            "  } catch(e){}\n"
            "})();\n</script>"
        )
        html = html.replace("</body>", personalise_js + "\n</body>")

    # Inject home button (fixed position, always visible)
    if home_url:
        home_btn = (
            f'<a href="{home_url}" style="position:fixed;top:12px;right:12px;z-index:9999;'
            'background:#1565C0;color:#fff;padding:7px 16px;border-radius:20px;'
            'text-decoration:none;font-size:12px;font-weight:700;'
            'box-shadow:0 2px 8px rgba(0,0,0,0.2);font-family:-apple-system,sans-serif">'
            '← Home</a>'
        )
        html = html.replace("</body>", home_btn + "\n</body>")

    out = html_path or HTML_PATH
    out.write_text(html, encoding="utf-8")
    print(f"  → {out.name} written")


# ── Main ──────────────────────────────────────────────────────────────────────


# ── Galas tab ─────────────────────────────────────────────────────────────────
def load_galas():
    """Load galas.json. Returns list of gala dicts, or [] if not found."""
    import json as _j
    if not GALAS_PATH.exists():
        return []
    try:
        return _j.loads(GALAS_PATH.read_text())
    except Exception:
        return []


def build_galas_html(galas, all_bests):
    """
    Build galas tab HTML.
    all_bests: {name: {(event, course): {time_str, time_secs, date_str}}}
    """
    from datetime import date as _date, datetime as _dt

    if not galas:
        return '<p style="color:#888;font-size:13px;padding:20px 0">No upcoming galas added yet.</p>'

    today = _date.today()
    blocks = []

    for gala in galas:
        name     = gala.get("name", "Unnamed Gala")
        course   = gala.get("course", "SC")
        venue    = gala.get("venue", "")
        raw_date = gala.get("date", "TBD")
        entries  = gala.get("entries", {})

        gala_date = None
        if raw_date and raw_date != "TBD":
            try:
                gala_date = _dt.strptime(raw_date, "%Y-%m-%d").date()
            except ValueError:
                pass

        if gala_date:
            days_diff = (gala_date - today).days
            past = days_diff < 0
            if days_diff < 0:
                countdown = '<span style="color:#888">' + str(abs(days_diff)) + 'd ago</span>'
            elif days_diff == 0:
                countdown = '<span style="color:#e74c3c;font-weight:700">Today!</span>'
            else:
                countdown = '<span style="color:#27ae60;font-weight:600">in ' + str(days_diff) + 'd</span>'
            date_str = gala_date.strftime("%-d %b %Y")
        else:
            countdown = '<span style="color:#888">Date TBC</span>'
            date_str = "TBC"
            past = False

        all_events = []
        seen_evs = set()
        for evs in entries.values():
            for ev in evs:
                if ev not in seen_evs:
                    all_events.append(ev)
                    seen_evs.add(ev)

        swimmers_in_gala = list(entries.keys())

        opacity_style = "opacity:0.6;" if past else ""
        course_bg = "#2e75b6" if course == "LC" else "#27ae60"
        venue_html = ""
        if venue and venue not in ("TBD", ""):
            venue_html = '<span style="font-size:12px;color:#888">\U0001f4cd ' + venue + '</span>'

        header = (
            '<div class="gala-card" style="' + opacity_style + 'background:white;border-radius:8px;'
            'box-shadow:0 1px 4px rgba(0,0,0,.1);padding:16px;margin-bottom:20px">'
            '<div style="display:flex;align-items:baseline;gap:12px;flex-wrap:wrap;margin-bottom:10px">'
            '<h2 style="margin:0;font-size:15px;color:#1a3a5c">' + name + '</h2>'
            '<span style="font-size:12px;background:' + course_bg + ';color:white;'
            'padding:1px 8px;border-radius:10px;font-weight:600">' + course + '</span>'
            '<span style="font-size:12px;color:#555">' + date_str + '</span>'
            '<span style="font-size:12px">' + countdown + '</span>'
            + venue_html +
            '</div>'
        )
        blocks.append(header)

        if not all_events or not swimmers_in_gala:
            blocks.append('<p style="color:#888;font-size:12px">No entries yet.</p></div>')
            continue

        # Find fastest gala-course PB per event for highlighting
        fastest = {}
        for ev in all_events:
            min_secs = None
            for sw in swimmers_in_gala:
                if ev not in entries.get(sw, []):
                    continue
                pb = all_bests.get(sw, {}).get((ev, course))
                if pb and pb.get("time_secs"):
                    if min_secs is None or pb["time_secs"] < min_secs:
                        min_secs = pb["time_secs"]
            fastest[ev] = min_secs

        sw_headers = ""
        for sw in swimmers_in_gala:
            parts = sw.split()
            first = parts[0] if parts else sw
            last  = parts[-1] if len(parts) > 1 else ""
            sw_headers += (
                '<th style="background:#1f4e79;color:white;font-size:11px;font-weight:600;'
                'padding:6px 10px;min-width:130px">' + first + '<br>'
                '<span style="font-weight:400;font-size:10px">' + last + '</span></th>'
            )

        def _pb_cell_html(pb, label, label_color):
            """Render one course row (SC or LC) inside a cell."""
            lbl = '<span style="font-size:9px;font-weight:700;color:' + label_color + '">' + label + '</span>'
            if pb and pb.get("time_str"):
                time_html = '<span style="font-weight:700;font-size:12px">' + pb["time_str"] + '</span>'
                date_html = ('<br><span style="font-size:9px;color:#999">' + pb["date_str"] + '</span>') if pb.get("date_str") else ''
                return lbl + ' ' + time_html + date_html
            else:
                return lbl + ' <span style="color:#ccc;font-size:11px">—</span>'

        rows_html = ""
        for ev in all_events:
            row = ('<tr><td style="font-weight:600;font-size:12px;padding:6px 10px;border:1px solid #eee;'
                   'background:#f8f9fa;white-space:nowrap">' + ev + '</td>')
            for sw in swimmers_in_gala:
                if ev not in entries.get(sw, []):
                    row += '<td style="color:#ddd;font-size:11px;text-align:center;padding:6px 8px;border:1px solid #eee">—</td>'
                    continue
                bests = all_bests.get(sw, {})
                sc_pb = bests.get((ev, "SC"))
                lc_pb = bests.get((ev, "LC"))
                gala_pb = bests.get((ev, course))
                is_fastest = (
                    gala_pb and gala_pb.get("time_secs") and
                    fastest.get(ev) is not None and
                    abs(gala_pb["time_secs"] - fastest[ev]) < 0.001
                )
                bg = "background:#d4edda;" if is_fastest else ""
                sc_html = _pb_cell_html(sc_pb, "SC", "#27ae60")
                lc_html = _pb_cell_html(lc_pb, "LC", "#2e75b6")
                divider = '<hr style="margin:4px 0;border:0;border-top:1px solid #eee">'
                row += ('<td style="' + bg + 'text-align:center;padding:6px 8px;border:1px solid #eee;'
                        'vertical-align:middle">' + sc_html + divider + lc_html + '</td>')
            row += '</tr>'
            rows_html += row

        event_th = ('<th style="background:#1f4e79;color:white;font-size:11px;font-weight:600;'
                    'padding:6px 10px;text-align:left;min-width:120px">Event</th>')
        table = ('<div style="overflow-x:auto"><table style="border-collapse:collapse;width:100%">'
                 '<thead><tr>' + event_th + sw_headers + '</tr></thead>'
                 '<tbody>' + rows_html + '</tbody>'
                 '</table></div></div>')
        blocks.append(table)

    return "\n".join(blocks)


def main():
    import sys
    refresh      = "--refresh" in sys.argv       # legacy: refreshes all histories (BSC + Ava)
    fetch_history = "--fetch-history" in sys.argv # BSC only: fetch/update all BSC swimmer histories

    run_time = datetime.now().strftime("%d/%m/%Y %H:%M")
    print(f"Swimming Results — {run_time}")
    _mode = "FETCH HISTORY (all BSC swimmers)" if fetch_history else ("FULL REFRESH" if refresh else "standard (using cached peer histories)")
    print(f"Mode: {_mode}")
    print("=" * 50)

    # ── Margot — always first ─────────────────────────────────────────────────
    print(f"\nFetching Margot's club PBs...")
    swimmers = fetch_all_swimmers()

    # Patch Margot's bests with any manual swims faster than Swim England's current PB
    _manual_path = MARGOT_DIR / "margot_manual_swims.json"
    if _manual_path.exists():
        import json as _mj
        _manual = _mj.loads(_manual_path.read_text())
        _margot = next((s for s in swimmers if s["is_margot"]), None)
        if _margot:
            for _ekey, _swims in _manual.items():
                _ev, _co = _ekey.split('|')
                for _s in _swims:
                    _secs = time_to_seconds(_s['time'])
                    if not _secs:
                        continue
                    _cur = _margot['bests'].get((_ev, _co), {})
                    if not _cur or _secs < _cur.get('time_secs', float('inf')):
                        _margot['bests'][(_ev, _co)] = {
                            'time_str':  _s['time'],
                            'time_secs': _secs,
                            'date_str':  _s['date'],
                            'converted': '',
                        }

    print(f"\nBuilding Margot's comparison matrix...")
    rows = build_matrix(swimmers)

    print(f"\nFetching Margot's swim history...")
    history = fetch_margot_history()
    # Merge manual swims into history so progress arrows and history modal both see them
    if _manual_path.exists():
        import json as _hj
        for _hkey, _hswims in _hj.loads(_manual_path.read_text()).items():
            _hev, _hco = _hkey.split('|')
            _hexist = history.setdefault((_hev, _hco), [])
            _hseen = {(s['date'], s['time']) for s in _hexist}
            for _hs in _hswims:
                if (_hs['date'], _hs['time']) not in _hseen:
                    _hexist.append(_hs)
    print(f"  → {sum(len(v) for v in history.values())} swims across {len(history)} events")

    print(f"\nReading qualifying times and previous bests...")
    quals      = read_margot_quals()
    prev_bests = load_previous_bests()

    update_pbs(swimmers)
    save_current_bests(swimmers)

    # Peer histories — fetch on --fetch-history or --refresh, otherwise load cache
    if refresh or fetch_history:
        print(f"\nFetching histories for all BSC 2014 swimmers...")
        peer_histories = fetch_peer_histories(swimmers, PEER_HISTORY_PATH, MARGOT_NAME)
    else:
        peer_histories = load_peer_histories(PEER_HISTORY_PATH)
        if peer_histories:
            print(f"\nLoaded cached peer histories ({len(peer_histories)} swimmers)")
        else:
            print(f"\nNo peer history cache — run with --fetch-history to populate")

    # Fetch all BSC peer age groups (2013–2018 except primary 2014)
    _bsc_peers = [
        (2013, CONFIG_2013_PATH, PEER_HIST_2013_PATH),
        (2015, CONFIG_2015_PATH, PEER_HIST_2015_PATH),
        (2016, CONFIG_2016_PATH, PEER_HIST_2016_PATH),
        (2017, CONFIG_2017_PATH, PEER_HIST_2017_PATH),
        (2018, CONFIG_2018_PATH, PEER_HIST_2018_PATH),
    ]
    margot_peer_rows_by_yob = {}
    all_peer_histories = {**(peer_histories or {})}
    age_group_quals = {}
    _all_margot_peer_swimmers = []
    for _yob, _cfg, _hist in _bsc_peers:
        if not _cfg.exists():
            continue
        print(f"\nFetching BSC {2026 - _yob}yo ({_yob}) PBs...")
        _sws = fetch_peer_swimmers(_cfg, _yob)
        if refresh or fetch_history:
            print(f"\nFetching BSC {2026 - _yob}yo histories...")
            _ph = fetch_peer_histories(_sws, _hist, "")
        else:
            _ph = load_peer_histories(_hist) or {}
            if _ph:
                print(f"\nLoaded cached {2026 - _yob}yo histories ({len(_ph)} swimmers)")
        margot_peer_rows_by_yob[_yob] = build_peer_rows(_sws)
        _all_margot_peer_swimmers += list(_sws)
        all_peer_histories.update(_ph or {})
        _aq = read_age_group_quals(_yob)
        if _aq:
            age_group_quals[_yob] = _aq

    build_html(rows, run_time, history=history, quals=quals,
               prev_bests=prev_bests, peer_histories=all_peer_histories,
               peer_rows_by_yob=margot_peer_rows_by_yob,
               age_group_quals=age_group_quals,
               html_path=HTML_PATH, home_url="brompton_home.html?from=personal",
               manifest="bsc_manifest.json")


    # ── Ava ───────────────────────────────────────────────────────────────────
    print(f"\nFetching Ava's club PBs...")
    ava_swimmers = fetch_ava_club_swimmers()

    # Patch Ava's bests with any manual swims faster than Swim England's current PB
    _ava_manual_path = AVA_DIR / "ava_manual_swims.json"
    if _ava_manual_path.exists():
        import json as _aj
        _ava_manual = _aj.loads(_ava_manual_path.read_text())
        _ava = next((s for s in ava_swimmers if s.get("is_margot")), None)
        if _ava:
            for _ekey, _swims in _ava_manual.items():
                _ev, _co = _ekey.split('|')
                for _s in _swims:
                    _secs = time_to_seconds(_s['time'])
                    if not _secs: continue
                    _cur = _ava['bests'].get((_ev, _co), {})
                    if not _cur or _secs < _cur.get('time_secs', float('inf')):
                        _ava['bests'][(_ev, _co)] = {'time_str': _s['time'], 'time_secs': _secs,
                                                     'date_str': _s['date'], 'converted': ''}

    print(f"\nBuilding Ava's comparison matrix...")
    ava_rows = build_matrix(ava_swimmers)

    print(f"\nFetching Ava's swim history...")
    ava_history = fetch_ava_history()
    print(f"  → {sum(len(v) for v in ava_history.values())} history swims for Ava")

    if refresh:
        print(f"\nFetching peer histories for all Ava's club swimmers...")
        ava_peer_histories = fetch_peer_histories(ava_swimmers, AVA_PEER_HIST_PATH, AVA_NAME)
    else:
        ava_peer_histories = load_peer_histories(AVA_PEER_HIST_PATH)
        if ava_peer_histories:
            print(f"\nLoaded cached Ava peer histories ({len(ava_peer_histories)} swimmers)")

    # ── Ava peer age groups (2013–2018 except primary 2017) ──────────────────
    _cwsc_peers = [
        (2013, AVA_CONFIG_2013_PATH, AVA_PEER_HIST_2013_PATH),
        (2014, AVA_CONFIG_2014_PATH, AVA_PEER_HIST_2014_PATH),
        (2015, AVA_CONFIG_2015_PATH, AVA_PEER_HIST_2015_PATH),
        (2016, AVA_CONFIG_2016_PATH, AVA_PEER_HIST_2016_PATH),
        (2018, AVA_CONFIG_2018_PATH, AVA_PEER_HIST_2018_PATH),
    ]
    ava_peer_rows_by_yob = {}
    all_ava_peer_histories = {**(ava_peer_histories or {})}
    ava_age_group_quals = {}
    _all_ava_peer_swimmers = []
    for _yob, _cfg, _hist in _cwsc_peers:
        if not _cfg.exists():
            continue
        print(f"\nFetching CWSC {2026 - _yob}yo ({_yob}) PBs...")
        _sws = fetch_peer_swimmers(_cfg, _yob)
        if refresh or not _hist.exists():
            print(f"\nFetching CWSC {2026 - _yob}yo histories...")
            _ph = fetch_peer_histories(_sws, _hist, "")
        else:
            _ph = load_peer_histories(_hist) or {}
            if _ph:
                print(f"\nLoaded cached {2026 - _yob}yo histories ({len(_ph)} swimmers)")
        ava_peer_rows_by_yob[_yob] = build_peer_rows(_sws)
        _all_ava_peer_swimmers += list(_sws)
        all_ava_peer_histories.update(_ph or {})
        _aq = read_age_group_quals(_yob)
        if _aq:
            ava_age_group_quals[_yob] = _aq

    build_ava_xlsx(ava_swimmers)
    ava_quals = read_ava_quals()
    build_html(ava_rows, run_time,
               history=ava_history, quals=ava_quals, prev_bests=None,
               html_path=AVA_HTML_PATH,
               title="CWSC — Swim Results Tracker",
               peer_histories=all_ava_peer_histories,
               peer_rows_by_yob=ava_peer_rows_by_yob,
               age_group_quals=ava_age_group_quals,
               home_url="cwsc.html",
               manifest="cwsc_u14_manifest.json")

    # ── Inject Galas tab — separate per page ─────────────────────────────────
    def _bests_lookup(group):
        out = {}
        for _sw in group:
            _n = _sw.get("name", "")
            if not _n:
                continue
            out[_n] = {
                k: {"time_str":  v.get("time_str", ""),
                    "time_secs": v.get("time_secs"),
                    "date_str":  v.get("date_str", "")}
                for k, v in _sw.get("bests", {}).items()
            }
        return out

    def _inject_galas(html_path, page_galas, bests):
        if not page_galas:
            return
        import re as _re, json as _json
        galas_html = build_galas_html(page_galas, bests)

        # ── Tab nav (above settings panel) ──────────────────────────────────
        _tab_nav = (
            '<div id="tab-nav" style="display:flex;gap:0;margin-bottom:14px;border-bottom:2px solid #1f4e79">'
            '<button id="tab-rankings-btn" onclick="switchTab(\'rankings\')" '
            'style="padding:8px 20px;background:#1f4e79;color:white;border:none;cursor:pointer;'
            'font-size:13px;font-weight:600;border-radius:4px 0 0 0;touch-action:manipulation">Rankings</button>'
            '<button id="tab-galas-btn" onclick="switchTab(\'galas\')" '
            'style="padding:8px 20px;background:#f0f4f8;color:#1f4e79;border:none;cursor:pointer;'
            'font-size:13px;font-weight:600;border-radius:0 4px 0 0;touch-action:manipulation">Galas</button>'
            '</div>'
        )

        # ── Gala filter dropdown (injected inside settings panel) ────────────
        _gala_opts = '<option value="">— All events —</option>'
        _sorted_galas = sorted(page_galas, key=lambda g: g.get("date", ""), reverse=True)
        for g in _sorted_galas:
            _safe = g['name'].replace('"', '&quot;')
            _raw_d = g.get("date", "")
            try:
                from datetime import datetime as _dtt
                _dlabel = _dtt.strptime(_raw_d, "%Y-%m-%d").strftime("%-d %b %Y")
            except Exception:
                _dlabel = _raw_d
            _gala_opts += '<option value="' + _safe + '">' + g['name'] + ' — ' + _dlabel + '</option>'
        _gala_filter_section = (
            '<div class="settings-section" id="gala-filter-section">'
            '<h3>Gala filter</h3>'
            '<select id="gala-select" onchange="applyGalaFilter(this.value)" '
            'style="font-size:11px;padding:3px 6px;border:1px solid #ccc;border-radius:4px;'
            'cursor:pointer;touch-action:manipulation">'
            + _gala_opts +
            '</select>'
            '<div style="margin-top:5px"><button onclick="applyGalaFilter(\'\');document.getElementById(\'gala-select\').value=\'\'" '
            'style="font-size:10px;padding:2px 8px;border:1px solid #ccc;border-radius:3px;background:#f9f9f9;cursor:pointer">Clear</button></div>'
            '</div>'
        )

        # ── Galas section (body) ─────────────────────────────────────────────
        _galas_div = '<div id="galas-section" style="display:none;padding:8px 0">' + galas_html + '</div>'

        # ── Tab switcher JS ──────────────────────────────────────────────────
        _tab_js = (
            "<script id='swim-tab-js'>\n"
            "window.switchTab=function(tab){\n"
            "  var isGalas=tab==='galas';\n"
            "  var gs=document.getElementById('galas-section');if(gs)gs.style.display=isGalas?'block':'none';\n"
            "  var s=document.querySelector('.settings');if(s)s.style.display=isGalas?'none':'';\n"
            "  var w=document.querySelector('.wrapper');if(w)w.style.display=isGalas?'none':'';\n"
            "  var rb=document.getElementById('tab-rankings-btn');\n"
            "  var gb=document.getElementById('tab-galas-btn');\n"
            "  if(rb){rb.style.background=isGalas?'#f0f4f8':'#1f4e79';rb.style.color=isGalas?'#1f4e79':'white';}\n"
            "  if(gb){gb.style.background=isGalas?'#1f4e79':'#f0f4f8';gb.style.color=isGalas?'white':'#1f4e79';}\n"
            "  try{localStorage.setItem('activeTab',tab);}catch(e){}\n"
            "};\n"
            "document.addEventListener('DOMContentLoaded',function(){\n"
            "  try{var t=localStorage.getItem('activeTab');if(t)switchTab(t);}catch(e){}\n"
            "});\n"
            "</script>"
        )

        # ── Gala filter JS ───────────────────────────────────────────────────
        _galas_js_data = _json.dumps(page_galas, ensure_ascii=False)
        _gala_filter_js = (
            "<script id='gala-filter-js'>\n"
            "var GALAS_DATA=" + _galas_js_data + ";\n"
            "function _galaSlug(s){return s.toLowerCase().replace(/ /g,'-').replace(/→/g,'-');}\n"
            "function applyGalaFilter(galaName){\n"
            "  if(!galaName){\n"
            "    document.querySelectorAll('.family-cb,.dist-cb').forEach(function(cb){cb.checked=true;});\n"
            "    document.querySelectorAll('.coltype-cb').forEach(function(cb){cb.checked=true;});\n"
            "  } else {\n"
            "    var gala=GALAS_DATA.find(function(g){return g.name===galaName;});\n"
            "    if(!gala)return;\n"
            "    var allEvs=[];\n"
            "    Object.values(gala.entries).forEach(function(evs){allEvs=allEvs.concat(evs);});\n"
            "    var evSlugs=allEvs.map(function(e){return _galaSlug(e);});\n"
            "    var activeFamilies={},activeDists={};\n"
            "    evSlugs.forEach(function(s){\n"
            "      var parts=s.split('-'),dist=parseInt(parts[0]),fam=parts.slice(1).join('-');\n"
            "      activeFamilies[fam]=true;\n"
            "      activeDists[dist<=50?'50m':(dist<=100?'100m':'200mp')]=true;\n"
            "    });\n"
            "    document.querySelectorAll('.family-cb').forEach(function(cb){\n"
            "      cb.checked=!!activeFamilies[cb.dataset.family];\n"
            "    });\n"
            "    document.querySelectorAll('.dist-cb').forEach(function(cb){\n"
            "      cb.checked=!!activeDists[cb.dataset.dist];\n"
            "    });\n"
            "    var c=gala.course;\n"
            "    document.querySelectorAll('.coltype-cb').forEach(function(cb){\n"
            "      var ct=cb.dataset.coltype;\n"
            "      if(c==='SC') cb.checked=(ct==='sc'||ct==='best');\n"
            "      else if(c==='LC') cb.checked=(ct==='lc'||ct==='best');\n"
            "      else cb.checked=true;\n"
            "    });\n"
            "  }\n"
            "  var _cb=document.querySelector('.family-cb');\n"
            "  if(_cb)_cb.dispatchEvent(new Event('change'));\n"
            "}\n"
            "</script>"
        )

        # ── Read and idempotently strip previous injections ──────────────────
        _txt = html_path.read_text(encoding="utf-8")
        _txt = _re.sub(r'<!-- SWIM-TAB-NAV-START -->.*?<!-- SWIM-TAB-NAV-END -->\n?', '', _txt, flags=_re.DOTALL)
        _txt = _re.sub(r'<!-- SWIM-GALA-FILTER-START -->.*?<!-- SWIM-GALA-FILTER-END -->\n?', '', _txt, flags=_re.DOTALL)
        _txt = _re.sub(r'<!-- SWIM-BODY-INJECT-START -->.*?<!-- SWIM-BODY-INJECT-END -->\n?', '', _txt, flags=_re.DOTALL)
        # Also strip old-style injection (from previous run format) for backward compat
        if 'id="tab-nav"' in _txt:
            _txt = _re.sub(r'<div id="tab-nav"[^>]*>.*?(?=<div class="settings")', '', _txt, flags=_re.DOTALL)
        if "<script id='swim-tab-js'>" in _txt:
            _txt = _re.sub(r"<script id='swim-tab-js'>.*?</script>\s*", '', _txt, flags=_re.DOTALL)

        # ── Inject tab nav before settings panel ────────────────────────────
        _tab_nav_block = '<!-- SWIM-TAB-NAV-START -->\n' + _tab_nav + '\n<!-- SWIM-TAB-NAV-END -->'
        _txt = _txt.replace('<div class="settings">', _tab_nav_block + '\n<div class="settings">', 1)

        # ── Inject gala filter before Save-as-default button ────────────────
        _save_btn_marker = '<div class="settings-section" style="justify-content:flex-end'
        _gala_block = '<!-- SWIM-GALA-FILTER-START -->\n' + _gala_filter_section + '\n<!-- SWIM-GALA-FILTER-END -->\n    '
        _txt = _txt.replace(_save_btn_marker, _gala_block + _save_btn_marker, 1)

        # ── Inject galas section + JS before </body> ─────────────────────────
        _body_block = (
            '<!-- SWIM-BODY-INJECT-START -->\n'
            + _galas_div + '\n'
            + _tab_js + '\n'
            + _gala_filter_js + '\n'
            + '<!-- SWIM-BODY-INJECT-END -->'
        )
        _txt = _txt.replace("</body>", _body_block + "\n</body>", 1)

        html_path.write_text(_txt, encoding="utf-8")
        print(f"  → Galas tab injected into {html_path.name}")


    # Galas tab has been removed from both U14 personal pages — do not re-inject


    # Ensure icon/manifest tags survive galas injection on both pages
    for _hp, _mf in [(HTML_PATH, "bsc_manifest.json"), (AVA_HTML_PATH, "cwsc_u14_manifest.json")]:
        _ht = _hp.read_text(encoding="utf-8")
        if 'apple-touch-icon' not in _ht:
            _icon_tags = f'\n<link rel="manifest" href="{_mf}">\n<link rel="apple-touch-icon" href="icon-192.png">'
            _ht = _ht.replace('</title>', '</title>' + _icon_tags, 1)
            _hp.write_text(_ht, encoding="utf-8")

    print(f"\nDone.")
    print(f"  Spreadsheet  : {XLSX_PATH.name}")
    print(f"  HTML report  : {HTML_PATH.name}")
    print(f"  Ava report   : {AVA_HTML_PATH.name}")
    print(f"  Ava workbook : {AVA_XLSX_PATH.name}")
    if refresh or fetch_history:
        print(f"  BSC histories cached for next run")


if __name__ == "__main__":
    main()
